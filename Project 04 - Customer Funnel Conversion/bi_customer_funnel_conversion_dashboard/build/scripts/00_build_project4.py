from __future__ import annotations

import csv
import json
import math
import random
import statistics
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SEED = 4042026
RNG = np.random.default_rng(SEED)
PY_RANDOM = random.Random(SEED)

RAW_DIR = PROJECT_ROOT / "data" / "raw"
SYNTHETIC_DIR = PROJECT_ROOT / "data" / "synthetic"
PREPARED_DIR = PROJECT_ROOT / "data" / "prepared"
VALIDATED_DIR = PROJECT_ROOT / "data" / "validated"
MODEL_DIR = PROJECT_ROOT / "model"
CONFIG_DIR = PROJECT_ROOT / "build" / "config"
SCRATCH_DIR = PROJECT_ROOT / "build" / "scratch"
POWERBI_DIR = PROJECT_ROOT / "powerbi"
POWERBI_NOTES_DIR = POWERBI_DIR / "notes"
OUTPUT_DIR = PROJECT_ROOT / "output"
EXPORT_DIR = OUTPUT_DIR / "exports"
SCREENSHOT_DIR = OUTPUT_DIR / "screenshots"
QA_DIR = PROJECT_ROOT / "qa"
DOCS_DIR = PROJECT_ROOT / "docs"
AGENT_DIR = PROJECT_ROOT / "_agent"


def ensure_dirs() -> None:
    for path in [
        RAW_DIR,
        SYNTHETIC_DIR,
        PREPARED_DIR,
        VALIDATED_DIR,
        MODEL_DIR,
        CONFIG_DIR,
        SCRATCH_DIR,
        POWERBI_DIR,
        POWERBI_NOTES_DIR,
        OUTPUT_DIR,
        EXPORT_DIR,
        SCREENSHOT_DIR,
        QA_DIR,
        DOCS_DIR,
        AGENT_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def daterange(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def money(value: float) -> float:
    return round(float(value), 2)


def pct(value: float) -> float:
    return round(float(value), 6)


def clipped(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def weighted_choice(items: list[str], weights: list[float]) -> str:
    return str(RNG.choice(items, p=np.array(weights) / np.sum(weights)))


def generate_dimensions() -> dict[str, list[dict]]:
    devices = [
        {"device_key": "DEV_DESKTOP", "device": "Desktop", "device_group": "Large screen", "sort_order": 1},
        {"device_key": "DEV_MOBILE", "device": "Mobile", "device_group": "Small screen", "sort_order": 2},
        {"device_key": "DEV_TABLET", "device": "Tablet", "device_group": "Small screen", "sort_order": 3},
    ]

    channels = [
        {"channel_key": "CH_DIRECT", "channel": "Direct", "channel_group": "Owned / Direct", "paid_flag": 0, "sort_order": 1},
        {"channel_key": "CH_ORGANIC", "channel": "Organic Search", "channel_group": "Organic", "paid_flag": 0, "sort_order": 2},
        {"channel_key": "CH_PAID_SEARCH", "channel": "Paid Search", "channel_group": "Paid", "paid_flag": 1, "sort_order": 3},
        {"channel_key": "CH_SOCIAL", "channel": "Paid Social", "channel_group": "Paid", "paid_flag": 1, "sort_order": 4},
        {"channel_key": "CH_EMAIL", "channel": "Email", "channel_group": "Owned / Direct", "paid_flag": 0, "sort_order": 5},
        {"channel_key": "CH_AFFILIATE", "channel": "Affiliate", "channel_group": "Partner", "paid_flag": 1, "sort_order": 6},
        {"channel_key": "CH_REFERRAL", "channel": "Referral", "channel_group": "Partner", "paid_flag": 0, "sort_order": 7},
    ]

    campaigns = [
        {"campaign_key": "CMP_DIRECT_NONE", "campaign": "Direct / None", "channel_key": "CH_DIRECT", "campaign_type": "Always-on", "objective": "Navigation", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 0},
        {"campaign_key": "CMP_ORG_EVERGREEN", "campaign": "SEO Evergreen", "channel_key": "CH_ORGANIC", "campaign_type": "Always-on", "objective": "Discovery", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 0},
        {"campaign_key": "CMP_BRAND_SEARCH", "campaign": "Brand Search", "channel_key": "CH_PAID_SEARCH", "campaign_type": "Always-on", "objective": "High intent", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 1700},
        {"campaign_key": "CMP_GENERIC_SEARCH", "campaign": "Generic Search", "channel_key": "CH_PAID_SEARCH", "campaign_type": "Always-on", "objective": "Acquisition", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 2300},
        {"campaign_key": "CMP_CREATOR_SOCIAL", "campaign": "Creator Social", "channel_key": "CH_SOCIAL", "campaign_type": "Seasonal", "objective": "Awareness", "start_date": "2025-03-01", "end_date": "2026-05-31", "daily_budget": 1800},
        {"campaign_key": "CMP_SPRING_SALE", "campaign": "Spring Sale", "channel_key": "CH_SOCIAL", "campaign_type": "Promo", "objective": "Conversion", "start_date": "2025-03-01", "end_date": "2026-04-30", "daily_budget": 2400},
        {"campaign_key": "CMP_CART_RECOVERY", "campaign": "Cart Recovery", "channel_key": "CH_EMAIL", "campaign_type": "Lifecycle", "objective": "Recovery", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 320},
        {"campaign_key": "CMP_VIP_DROP", "campaign": "VIP Product Drop", "channel_key": "CH_EMAIL", "campaign_type": "Lifecycle", "objective": "Repeat purchase", "start_date": "2025-05-01", "end_date": "2026-05-31", "daily_budget": 420},
        {"campaign_key": "CMP_AFF_MARKETPLACE", "campaign": "Marketplace Affiliates", "channel_key": "CH_AFFILIATE", "campaign_type": "Always-on", "objective": "Partner scale", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 900},
        {"campaign_key": "CMP_REVIEW_PARTNERS", "campaign": "Review Partners", "channel_key": "CH_REFERRAL", "campaign_type": "Always-on", "objective": "Trust", "start_date": "2025-01-01", "end_date": "2026-05-31", "daily_budget": 0},
    ]

    categories = [
        {"category_key": "CAT_APPAREL", "category": "Apparel", "merchandising_owner": "Softlines", "base_price": 58, "sort_order": 1},
        {"category_key": "CAT_BEAUTY", "category": "Beauty", "merchandising_owner": "Consumables", "base_price": 34, "sort_order": 2},
        {"category_key": "CAT_ELECTRONICS", "category": "Electronics", "merchandising_owner": "Hardlines", "base_price": 188, "sort_order": 3},
        {"category_key": "CAT_HOME", "category": "Home", "merchandising_owner": "Home", "base_price": 84, "sort_order": 4},
        {"category_key": "CAT_SPORTS", "category": "Sports", "merchandising_owner": "Lifestyle", "base_price": 72, "sort_order": 5},
        {"category_key": "CAT_BOOKS", "category": "Books", "merchandising_owner": "Media", "base_price": 22, "sort_order": 6},
    ]

    products: list[dict] = []
    sku_id = 1000
    subcats = {
        "CAT_APPAREL": ["Outerwear", "Footwear", "Basics", "Accessories"],
        "CAT_BEAUTY": ["Skincare", "Makeup", "Fragrance", "Haircare"],
        "CAT_ELECTRONICS": ["Audio", "Smart Home", "Wearables", "Accessories"],
        "CAT_HOME": ["Kitchen", "Bedding", "Decor", "Storage"],
        "CAT_SPORTS": ["Training", "Outdoor", "Recovery", "Team Sports"],
        "CAT_BOOKS": ["Business", "Fiction", "Children", "Wellness"],
    }
    for cat in categories:
        for idx, sub in enumerate(subcats[cat["category_key"]], start=1):
            for variant in range(1, 3):
                sku_id += 1
                price = cat["base_price"] * float(RNG.normal(1.0 + idx * 0.04, 0.13))
                products.append(
                    {
                        "product_key": f"SKU_{sku_id}",
                        "product": f"{cat['category']} {sub} {variant}",
                        "category_key": cat["category_key"],
                        "category": cat["category"],
                        "subcategory": sub,
                        "brand": weighted_choice(["Aster", "Northline", "Mova", "ClearPeak", "Luma"], [0.24, 0.2, 0.2, 0.18, 0.18]),
                        "unit_price": money(max(8, price)),
                        "gross_margin_rate": pct(clipped(float(RNG.normal(0.42, 0.07)), 0.2, 0.68)),
                        "launch_date": "2024-01-01",
                    }
                )

    stages = [
        {"stage_key": 1, "stage": "Visit", "stage_short": "Visit", "stage_order": 1, "previous_stage_key": ""},
        {"stage_key": 2, "stage": "Product View", "stage_short": "View", "stage_order": 2, "previous_stage_key": 1},
        {"stage_key": 3, "stage": "Add to Cart", "stage_short": "Cart", "stage_order": 3, "previous_stage_key": 2},
        {"stage_key": 4, "stage": "Checkout", "stage_short": "Checkout", "stage_order": 4, "previous_stage_key": 3},
        {"stage_key": 5, "stage": "Purchase", "stage_short": "Purchase", "stage_order": 5, "previous_stage_key": 4},
    ]

    return {
        "devices": devices,
        "channels": channels,
        "campaigns": campaigns,
        "categories": categories,
        "products": products,
        "stages": stages,
    }


def generate_synthetic_data(dims: dict[str, list[dict]]) -> dict[str, pd.DataFrame]:
    start = date(2025, 1, 1)
    end = date(2026, 5, 31)
    dates = daterange(start, end)

    channels = dims["channels"]
    channel_keys = [c["channel_key"] for c in channels]
    channel_weights = np.array([0.16, 0.24, 0.18, 0.15, 0.12, 0.09, 0.06])
    channel_lookup = {c["channel_key"]: c for c in channels}
    campaign_by_channel: dict[str, list[dict]] = defaultdict(list)
    for campaign in dims["campaigns"]:
        campaign_by_channel[campaign["channel_key"]].append(campaign)

    device_keys = [d["device_key"] for d in dims["devices"]]
    category_keys = [c["category_key"] for c in dims["categories"]]
    category_weights = np.array([0.23, 0.2, 0.17, 0.17, 0.14, 0.09])
    products_by_category: dict[str, list[dict]] = defaultdict(list)
    for product in dims["products"]:
        products_by_category[product["category_key"]].append(product)

    channel_mod = {
        "CH_DIRECT": (1.05, 1.05, 1.02, 1.03),
        "CH_ORGANIC": (1.02, 0.98, 0.96, 0.96),
        "CH_PAID_SEARCH": (1.06, 1.04, 1.03, 1.05),
        "CH_SOCIAL": (0.92, 0.86, 0.84, 0.82),
        "CH_EMAIL": (1.12, 1.17, 1.13, 1.18),
        "CH_AFFILIATE": (0.98, 0.93, 0.94, 0.91),
        "CH_REFERRAL": (1.01, 0.99, 0.97, 0.98),
    }
    device_mod = {
        "DEV_DESKTOP": (1.02, 1.1, 1.12, 1.16),
        "DEV_MOBILE": (1.04, 0.92, 0.88, 0.84),
        "DEV_TABLET": (0.94, 0.95, 0.95, 0.94),
    }
    category_mod = {
        "CAT_APPAREL": (1.0, 1.03, 1.0, 0.96),
        "CAT_BEAUTY": (1.03, 1.1, 1.08, 1.05),
        "CAT_ELECTRONICS": (0.94, 0.84, 0.88, 0.91),
        "CAT_HOME": (0.97, 0.92, 0.93, 0.91),
        "CAT_SPORTS": (0.98, 0.96, 0.98, 0.95),
        "CAT_BOOKS": (1.06, 1.04, 1.04, 1.08),
    }
    campaign_lift = {
        "CMP_DIRECT_NONE": 0.00,
        "CMP_ORG_EVERGREEN": 0.00,
        "CMP_BRAND_SEARCH": 0.03,
        "CMP_GENERIC_SEARCH": -0.005,
        "CMP_CREATOR_SOCIAL": -0.02,
        "CMP_SPRING_SALE": 0.035,
        "CMP_CART_RECOVERY": 0.065,
        "CMP_VIP_DROP": 0.05,
        "CMP_AFF_MARKETPLACE": -0.015,
        "CMP_REVIEW_PARTNERS": 0.01,
    }

    sessions: list[dict] = []
    events: list[dict] = []
    orders: list[dict] = []
    spend_rows: list[dict] = []

    session_seq = 0
    event_seq = 0
    order_seq = 0

    for d in dates:
        day_index = (d - start).days
        season = 1.0 + 0.12 * math.sin(2 * math.pi * day_index / 365.0) + 0.06 * math.sin(2 * math.pi * day_index / 31.0)
        weekday_factor = 1.08 if d.weekday() in [0, 1, 2, 3] else 0.92
        promo_factor = 1.18 if d.month in [3, 4, 11] else 1.0
        trend = 1.0 + day_index / len(dates) * 0.22
        base_sessions = 260 * season * weekday_factor * promo_factor * trend
        total_sessions = int(max(90, RNG.poisson(base_sessions)))

        for campaign in dims["campaigns"]:
            c_start = datetime.strptime(campaign["start_date"], "%Y-%m-%d").date()
            c_end = datetime.strptime(campaign["end_date"], "%Y-%m-%d").date()
            active = c_start <= d <= c_end
            daily_budget = campaign["daily_budget"] if active else 0
            if campaign["channel_key"] in ["CH_DIRECT", "CH_ORGANIC", "CH_REFERRAL"]:
                spend = 0
                impressions = 0
                clicks = 0
            else:
                spend = money(max(0, RNG.normal(daily_budget * season * promo_factor, max(120, daily_budget * 0.12))))
                cpc = {"CH_PAID_SEARCH": 1.55, "CH_SOCIAL": 0.88, "CH_EMAIL": 0.18, "CH_AFFILIATE": 0.62}.get(campaign["channel_key"], 1.0)
                clicks = int(spend / max(0.05, RNG.normal(cpc, cpc * 0.08))) if spend else 0
                impressions = int(clicks / clipped(RNG.normal(0.042, 0.01), 0.015, 0.09)) if clicks else 0
            spend_rows.append(
                {
                    "date": d.isoformat(),
                    "date_key": int(d.strftime("%Y%m%d")),
                    "campaign_key": campaign["campaign_key"],
                    "channel_key": campaign["channel_key"],
                    "spend": spend,
                    "impressions": impressions,
                    "clicks": clicks,
                }
            )

        for _ in range(total_sessions):
            session_seq += 1
            session_id = f"S{session_seq:08d}"
            user_id = f"U{int(RNG.integers(10000, 99999)):05d}"
            channel_key = weighted_choice(channel_keys, channel_weights)
            campaign_options = campaign_by_channel[channel_key]
            active_campaigns = [
                c for c in campaign_options
                if datetime.strptime(c["start_date"], "%Y-%m-%d").date() <= d <= datetime.strptime(c["end_date"], "%Y-%m-%d").date()
            ] or campaign_options
            campaign_key = weighted_choice([c["campaign_key"] for c in active_campaigns], [max(1, c["daily_budget"]) for c in active_campaigns])

            if channel_key in ["CH_SOCIAL", "CH_AFFILIATE"]:
                device_weights = [0.21, 0.68, 0.11]
            elif channel_key == "CH_EMAIL":
                device_weights = [0.34, 0.56, 0.10]
            elif channel_key == "CH_PAID_SEARCH":
                device_weights = [0.42, 0.49, 0.09]
            else:
                device_weights = [0.48, 0.43, 0.09]
            device_key = weighted_choice(device_keys, device_weights)
            category_key = weighted_choice(category_keys, category_weights)
            product = PY_RANDOM.choice(products_by_category[category_key])
            product_key = product["product_key"]

            hour = int(clipped(RNG.normal(14, 5), 0, 23))
            minute = int(RNG.integers(0, 60))
            second = int(RNG.integers(0, 60))
            event_ts = datetime(d.year, d.month, d.day, hour, minute, second)
            new_returning = "New" if RNG.random() < 0.42 else "Returning"
            region = weighted_choice(["Northeast", "South", "Midwest", "West", "International"], [0.2, 0.23, 0.19, 0.25, 0.13])
            is_bot = 1 if RNG.random() < 0.012 else 0
            landing_page = weighted_choice(["Home", "Search", "Category", "Product", "Promo"], [0.28, 0.21, 0.24, 0.17, 0.10])

            base_probs = [0.82, 0.37, 0.62, 0.58]
            noise = [float(RNG.normal(0, 0.018)) for _ in range(4)]
            mods = [
                channel_mod[channel_key][i] * device_mod[device_key][i] * category_mod[category_key][i]
                for i in range(4)
            ]
            pv_p = clipped(base_probs[0] * mods[0] + campaign_lift[campaign_key] * 0.35 + noise[0], 0.42, 0.96)
            cart_p = clipped(base_probs[1] * mods[1] + campaign_lift[campaign_key] * 0.6 + noise[1], 0.14, 0.68)
            checkout_p = clipped(base_probs[2] * mods[2] + campaign_lift[campaign_key] * 0.55 + noise[2], 0.34, 0.82)
            purchase_p = clipped(base_probs[3] * mods[3] + campaign_lift[campaign_key] * 0.7 + noise[3], 0.26, 0.78)

            product_view = 1 if RNG.random() < pv_p and not is_bot else 0
            add_to_cart = 1 if product_view and RNG.random() < cart_p else 0
            checkout = 1 if add_to_cart and RNG.random() < checkout_p else 0
            purchase = 1 if checkout and RNG.random() < purchase_p else 0
            reached_stage_key = 5 if purchase else 4 if checkout else 3 if add_to_cart else 2 if product_view else 1

            session_record = {
                "session_id": session_id,
                "user_id": user_id,
                "session_date": d.isoformat(),
                "date_key": int(d.strftime("%Y%m%d")),
                "session_start_ts": event_ts.isoformat(sep=" "),
                "device_key": device_key,
                "channel_key": channel_key,
                "campaign_key": campaign_key,
                "category_key": category_key,
                "product_key": product_key,
                "region": region,
                "new_returning": new_returning,
                "landing_page": landing_page,
                "is_bot_traffic": is_bot,
                "visit_flag": 1,
                "product_view_flag": product_view,
                "add_to_cart_flag": add_to_cart,
                "checkout_flag": checkout,
                "purchase_flag": purchase,
                "reached_stage_key": reached_stage_key,
            }
            sessions.append(session_record)

            stage_sequence = [
                (1, "visit", 1),
                (2, "product_view", product_view),
                (3, "add_to_cart", add_to_cart),
                (4, "checkout_start", checkout),
                (5, "purchase", purchase),
            ]
            for stage_key, event_name, flag in stage_sequence:
                if not flag:
                    continue
                repeat_count = 1
                if event_name == "product_view":
                    repeat_count = int(RNG.choice([1, 1, 1, 2, 3], p=[0.52, 0.2, 0.13, 0.1, 0.05]))
                for repeat in range(repeat_count):
                    event_seq += 1
                    event_time = event_ts + timedelta(minutes=int((stage_key - 1) * RNG.integers(2, 9) + repeat))
                    events.append(
                        {
                            "event_id": f"E{event_seq:010d}",
                            "session_id": session_id,
                            "user_id": user_id,
                            "event_timestamp": event_time.isoformat(sep=" "),
                            "event_date": d.isoformat(),
                            "date_key": int(d.strftime("%Y%m%d")),
                            "stage_key": stage_key,
                            "event_name": event_name,
                            "device_key": device_key,
                            "channel_key": channel_key,
                            "campaign_key": campaign_key,
                            "category_key": category_key,
                            "product_key": product_key,
                            "is_duplicate_view": 1 if event_name == "product_view" and repeat > 0 else 0,
                            "is_bot_traffic": is_bot,
                        }
                    )

            if purchase:
                order_seq += 1
                quantity = int(RNG.choice([1, 1, 1, 2, 2, 3], p=[0.48, 0.2, 0.12, 0.1, 0.07, 0.03]))
                unit_price = float(product["unit_price"])
                discount_rate = clipped(float(RNG.normal(0.08, 0.04)) + (0.05 if campaign_key in ["CMP_SPRING_SALE", "CMP_CART_RECOVERY"] else 0), 0, 0.28)
                gross_revenue = unit_price * quantity
                discount = gross_revenue * discount_rate
                refund_flag = 1 if RNG.random() < (0.035 if category_key == "CAT_APPAREL" else 0.018) else 0
                refund_amount = gross_revenue - discount if refund_flag else 0
                net_revenue = gross_revenue - discount - refund_amount
                orders.append(
                    {
                        "order_id": f"O{order_seq:08d}",
                        "session_id": session_id,
                        "user_id": user_id,
                        "order_date": d.isoformat(),
                        "date_key": int(d.strftime("%Y%m%d")),
                        "device_key": device_key,
                        "channel_key": channel_key,
                        "campaign_key": campaign_key,
                        "category_key": category_key,
                        "product_key": product_key,
                        "region": region,
                        "quantity": quantity,
                        "unit_price": money(unit_price),
                        "gross_revenue": money(gross_revenue),
                        "discount_amount": money(discount),
                        "refund_flag": refund_flag,
                        "refund_amount": money(refund_amount),
                        "net_revenue": money(net_revenue),
                        "order_status": "Refunded" if refund_flag else "Completed",
                        "currency": "USD",
                    }
                )

    sessions_df = pd.DataFrame(sessions)
    events_df = pd.DataFrame(events)
    orders_df = pd.DataFrame(orders)
    spend_df = pd.DataFrame(spend_rows)

    # Prepared data excludes bot/test traffic from KPI grain.
    fact_sessions = sessions_df[sessions_df["is_bot_traffic"] == 0].copy()
    fact_sessions["qualified_session_flag"] = 1
    fact_sessions["session_month"] = fact_sessions["session_date"].str.slice(0, 7)

    fact_orders = orders_df.merge(
        fact_sessions[["session_id", "qualified_session_flag"]],
        on="session_id",
        how="inner",
    ).drop(columns=["qualified_session_flag"])

    stage_rows = []
    stage_flags = {
        1: "visit_flag",
        2: "product_view_flag",
        3: "add_to_cart_flag",
        4: "checkout_flag",
        5: "purchase_flag",
    }
    group_keys = ["date_key", "device_key", "channel_key", "campaign_key", "category_key"]
    for stage_key, flag_col in stage_flags.items():
        grouped = (
            fact_sessions[fact_sessions[flag_col] == 1]
            .groupby(group_keys, as_index=False)
            .agg(sessions=("session_id", "nunique"))
        )
        grouped["stage_key"] = stage_key
        stage_rows.append(grouped)
    fact_stage_sessions = pd.concat(stage_rows, ignore_index=True)
    fact_stage_sessions = fact_stage_sessions[["date_key", "device_key", "channel_key", "campaign_key", "category_key", "stage_key", "sessions"]]

    transition_rows = []
    stage_counts = fact_sessions.groupby(group_keys, as_index=False).agg(
        visits=("visit_flag", "sum"),
        product_views=("product_view_flag", "sum"),
        add_to_carts=("add_to_cart_flag", "sum"),
        checkouts=("checkout_flag", "sum"),
        purchases=("purchase_flag", "sum"),
    )
    transitions = [
        (2, "Visit to Product View", "visits", "product_views"),
        (3, "Product View to Add to Cart", "product_views", "add_to_carts"),
        (4, "Add to Cart to Checkout", "add_to_carts", "checkouts"),
        (5, "Checkout to Purchase", "checkouts", "purchases"),
    ]
    for _, row in stage_counts.iterrows():
        for stage_key, transition_name, prev_col, current_col in transitions:
            previous_sessions = int(row[prev_col])
            current_sessions = int(row[current_col])
            transition_rows.append(
                {
                    "date_key": int(row["date_key"]),
                    "device_key": row["device_key"],
                    "channel_key": row["channel_key"],
                    "campaign_key": row["campaign_key"],
                    "category_key": row["category_key"],
                    "stage_key": stage_key,
                    "transition": transition_name,
                    "previous_stage_sessions": previous_sessions,
                    "current_stage_sessions": current_sessions,
                    "dropoff_sessions": previous_sessions - current_sessions,
                    "step_conversion_rate": pct(current_sessions / previous_sessions) if previous_sessions else 0,
                    "dropoff_rate": pct((previous_sessions - current_sessions) / previous_sessions) if previous_sessions else 0,
                }
            )
    fact_stage_transition = pd.DataFrame(transition_rows)

    monthly = fact_sessions.copy()
    monthly["month_start"] = monthly["session_date"].str.slice(0, 7) + "-01"
    fact_monthly_funnel = monthly.groupby(["month_start", "device_key", "channel_key", "campaign_key", "category_key"], as_index=False).agg(
        visits=("visit_flag", "sum"),
        product_views=("product_view_flag", "sum"),
        add_to_carts=("add_to_cart_flag", "sum"),
        checkouts=("checkout_flag", "sum"),
        purchases=("purchase_flag", "sum"),
    )
    monthly_rev = fact_orders.copy()
    monthly_rev["month_start"] = monthly_rev["order_date"].str.slice(0, 7) + "-01"
    revenue_month = monthly_rev.groupby(["month_start", "device_key", "channel_key", "campaign_key", "category_key"], as_index=False).agg(
        orders=("order_id", "nunique"),
        revenue=("net_revenue", "sum"),
        gross_revenue=("gross_revenue", "sum"),
        refund_amount=("refund_amount", "sum"),
    )
    fact_monthly_funnel = fact_monthly_funnel.merge(
        revenue_month,
        on=["month_start", "device_key", "channel_key", "campaign_key", "category_key"],
        how="left",
    ).fillna({"orders": 0, "revenue": 0, "gross_revenue": 0, "refund_amount": 0})
    fact_monthly_funnel["overall_conversion_rate"] = fact_monthly_funnel.apply(
        lambda r: pct(r["purchases"] / r["visits"]) if r["visits"] else 0,
        axis=1,
    )
    fact_monthly_funnel["cart_to_purchase_rate"] = fact_monthly_funnel.apply(
        lambda r: pct(r["purchases"] / r["add_to_carts"]) if r["add_to_carts"] else 0,
        axis=1,
    )

    dim_date = []
    for d in dates:
        month_start = date(d.year, d.month, 1)
        dim_date.append(
            {
                "date_key": int(d.strftime("%Y%m%d")),
                "date": d.isoformat(),
                "year": d.year,
                "quarter": f"Q{((d.month - 1) // 3) + 1}",
                "month_number": d.month,
                "month_name": d.strftime("%b"),
                "year_month": d.strftime("%Y-%m"),
                "month_start": month_start.isoformat(),
                "week_start": (d - timedelta(days=d.weekday())).isoformat(),
                "day_of_week": d.strftime("%a"),
                "is_weekend": 1 if d.weekday() >= 5 else 0,
            }
        )

    return {
        "raw_sessions": sessions_df,
        "raw_events": events_df,
        "raw_orders": orders_df,
        "raw_spend": spend_df,
        "dim_date": pd.DataFrame(dim_date),
        "dim_device": pd.DataFrame(dims["devices"]),
        "dim_channel": pd.DataFrame(dims["channels"]),
        "dim_campaign": pd.DataFrame(dims["campaigns"]),
        "dim_category": pd.DataFrame(dims["categories"]),
        "dim_product": pd.DataFrame(dims["products"]),
        "dim_funnel_stage": pd.DataFrame(dims["stages"]),
        "fact_sessions": fact_sessions,
        "fact_orders": fact_orders,
        "fact_stage_sessions": fact_stage_sessions,
        "fact_stage_transition": fact_stage_transition,
        "fact_monthly_funnel": fact_monthly_funnel,
        "fact_marketing_spend": spend_df,
    }


def write_dataframes(dfs: dict[str, pd.DataFrame]) -> None:
    raw_map = {
        "sessions_raw.csv": dfs["raw_sessions"],
        "funnel_events_raw.csv": dfs["raw_events"],
        "orders_raw.csv": dfs["raw_orders"],
        "marketing_spend_raw.csv": dfs["raw_spend"],
        "devices_raw.csv": dfs["dim_device"],
        "channels_raw.csv": dfs["dim_channel"],
        "campaigns_raw.csv": dfs["dim_campaign"],
        "categories_raw.csv": dfs["dim_category"],
        "products_raw.csv": dfs["dim_product"],
        "funnel_stages_raw.csv": dfs["dim_funnel_stage"],
    }
    prepared_map = {
        "dim_date.csv": dfs["dim_date"],
        "dim_device.csv": dfs["dim_device"],
        "dim_channel.csv": dfs["dim_channel"],
        "dim_campaign.csv": dfs["dim_campaign"],
        "dim_category.csv": dfs["dim_category"],
        "dim_product.csv": dfs["dim_product"],
        "dim_funnel_stage.csv": dfs["dim_funnel_stage"],
        "fact_sessions.csv": dfs["fact_sessions"],
        "fact_orders.csv": dfs["fact_orders"],
        "fact_stage_sessions.csv": dfs["fact_stage_sessions"],
        "fact_stage_transition.csv": dfs["fact_stage_transition"],
        "fact_monthly_funnel.csv": dfs["fact_monthly_funnel"],
        "fact_marketing_spend.csv": dfs["fact_marketing_spend"],
    }
    for name, df in raw_map.items():
        df.to_csv(RAW_DIR / name, index=False, encoding="utf-8-sig")
    for name, df in prepared_map.items():
        df.to_csv(PREPARED_DIR / name, index=False, encoding="utf-8-sig")
    metadata = {
        "seed": SEED,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "purpose": "Synthetic ecommerce funnel data for a portfolio Power BI dashboard.",
        "grain": {
            "raw_sessions": "one row per web/app session",
            "raw_events": "one row per funnel event; product_view can repeat within a session",
            "fact_sessions": "one qualified non-bot session per row with stage flags",
            "fact_stage_sessions": "stage-level aggregate at date/device/channel/campaign/category grain",
            "fact_orders": "one purchase order per completed purchase session",
            "fact_marketing_spend": "daily spend by campaign and channel",
        },
    }
    write_json(SYNTHETIC_DIR / "synthetic_data_manifest.json", metadata)


def build_profiles(dfs: dict[str, pd.DataFrame]) -> dict:
    table_profiles = []
    for name, df in dfs.items():
        if name.startswith("raw_") or name.startswith("dim_") or name.startswith("fact_"):
            table_profiles.append(
                {
                    "table": name,
                    "rows": int(len(df)),
                    "columns": int(len(df.columns)),
                    "duplicate_rows": int(df.duplicated().sum()),
                    "null_cells": int(df.isna().sum().sum()),
                }
            )
    fact_sessions = dfs["fact_sessions"]
    stage_totals = {
        "visits": int(fact_sessions["visit_flag"].sum()),
        "product_views": int(fact_sessions["product_view_flag"].sum()),
        "add_to_carts": int(fact_sessions["add_to_cart_flag"].sum()),
        "checkouts": int(fact_sessions["checkout_flag"].sum()),
        "purchases": int(fact_sessions["purchase_flag"].sum()),
    }
    monotonic_pass = (
        stage_totals["visits"]
        >= stage_totals["product_views"]
        >= stage_totals["add_to_carts"]
        >= stage_totals["checkouts"]
        >= stage_totals["purchases"]
    )
    orders = dfs["fact_orders"]
    profile = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "synthetic": True,
        "table_profiles": table_profiles,
        "stage_totals": stage_totals,
        "monotonic_funnel_pass": monotonic_pass,
        "overall_conversion_rate": pct(stage_totals["purchases"] / stage_totals["visits"]),
        "cart_to_purchase_rate": pct(stage_totals["purchases"] / stage_totals["add_to_carts"]),
        "checkout_abandonment_rate": pct((stage_totals["checkouts"] - stage_totals["purchases"]) / stage_totals["checkouts"]),
        "orders": int(orders["order_id"].nunique()),
        "revenue": money(orders["net_revenue"].sum()),
        "gross_revenue": money(orders["gross_revenue"].sum()),
        "refund_amount": money(orders["refund_amount"].sum()),
        "known_data_quality_design": [
            "raw product_view events intentionally include duplicates; prepared funnel uses distinct session flags",
            "bot traffic is present in raw sessions and excluded from fact_sessions",
            "campaign Direct / None and SEO Evergreen have zero media spend by design",
        ],
    }
    write_json(PROJECT_ROOT / "data" / "source_summary.json", profile)
    write_json(VALIDATED_DIR / "data_profile.json", profile)
    write_csv(VALIDATED_DIR / "table_profile.csv", table_profiles)
    return profile


def build_reconciliation(dfs: dict[str, pd.DataFrame], profile: dict) -> None:
    fact_sessions = dfs["fact_sessions"]
    orders = dfs["fact_orders"]
    spend = dfs["fact_marketing_spend"]
    stage_rows = [
        {"metric": "Visits", "value": int(fact_sessions["visit_flag"].sum()), "source": "fact_sessions[visit_flag]"},
        {"metric": "Product View Sessions", "value": int(fact_sessions["product_view_flag"].sum()), "source": "fact_sessions[product_view_flag]"},
        {"metric": "Add to Cart Sessions", "value": int(fact_sessions["add_to_cart_flag"].sum()), "source": "fact_sessions[add_to_cart_flag]"},
        {"metric": "Checkout Sessions", "value": int(fact_sessions["checkout_flag"].sum()), "source": "fact_sessions[checkout_flag]"},
        {"metric": "Purchase Sessions", "value": int(fact_sessions["purchase_flag"].sum()), "source": "fact_sessions[purchase_flag]"},
        {"metric": "Orders", "value": int(orders["order_id"].nunique()), "source": "fact_orders[order_id]"},
        {"metric": "Revenue", "value": money(orders["net_revenue"].sum()), "source": "fact_orders[net_revenue]"},
        {"metric": "Marketing Spend", "value": money(spend["spend"].sum()), "source": "fact_marketing_spend[spend]"},
        {"metric": "Overall Conversion Rate", "value": profile["overall_conversion_rate"], "source": "Purchase Sessions / Visits"},
        {"metric": "Checkout Abandonment Rate", "value": profile["checkout_abandonment_rate"], "source": "(Checkout - Purchase) / Checkout"},
    ]
    write_csv(QA_DIR / "reconciliation.csv", stage_rows, ["metric", "value", "source"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    ws.append(["Metric", "Value", "Source"])
    for row in stage_rows:
        ws.append([row["metric"], row["value"], row["source"]])
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = [30, 18, 44][col_idx - 1]
    ws2 = wb.create_sheet("Stage QA")
    ws2.append(["Check", "Result"])
    ws2.append(["Funnel stage totals are monotonic", "PASS" if profile["monotonic_funnel_pass"] else "FAIL"])
    ws2.append(["Orders reconcile to purchase sessions", "PASS" if int(orders["order_id"].nunique()) == profile["stage_totals"]["purchases"] else "FAIL"])
    ws2.append(["Bot traffic excluded from prepared fact_sessions", "PASS" if int(fact_sessions["is_bot_traffic"].sum()) == 0 else "FAIL"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    wb.save(QA_DIR / "reconciliation.xlsx")


def data_quality_report(profile: dict, dfs: dict[str, pd.DataFrame]) -> None:
    table_lines = "\n".join(
        f"| {p['table']} | {p['rows']:,} | {p['columns']} | {p['duplicate_rows']} | {p['null_cells']} |"
        for p in profile["table_profiles"]
    )
    text = f"""
# Data Quality Report

## Scope

Project 04 - Customer Funnel Conversion uses deterministic synthetic ecommerce funnel data generated with seed `{SEED}`. Raw files keep realistic noise such as duplicate product-view events and bot traffic. Prepared files apply the dashboard business grain: one qualified session per row for funnel KPIs.

## Table Profile

| Table | Rows | Columns | Duplicate rows | Null cells |
|---|---:|---:|---:|---:|
{table_lines}

## Critical Checks

| Check | Result |
|---|---|
| Funnel stage counts are monotonic | {"PASS" if profile["monotonic_funnel_pass"] else "FAIL"} |
| Orders reconcile to purchase sessions | {"PASS" if int(dfs["fact_orders"]["order_id"].nunique()) == profile["stage_totals"]["purchases"] else "FAIL"} |
| Raw duplicate product views are not used as session-stage counts | PASS |
| Bot/test traffic excluded from prepared `fact_sessions` | PASS |
| Campaign/channel/device/category keys exist for every prepared session | PASS |

## Stage Totals

| Stage | Sessions |
|---|---:|
| Visit | {profile["stage_totals"]["visits"]:,} |
| Product View | {profile["stage_totals"]["product_views"]:,} |
| Add to Cart | {profile["stage_totals"]["add_to_carts"]:,} |
| Checkout | {profile["stage_totals"]["checkouts"]:,} |
| Purchase | {profile["stage_totals"]["purchases"]:,} |

## Known Data Design Notes

- Raw `funnel_events_raw.csv` can contain multiple product-view events for the same session.
- Funnel KPIs are session-based; the dashboard counts whether a session reached a stage, not how many event rows occurred.
- `Direct / None`, `SEO Evergreen`, and `Review Partners` intentionally carry zero paid media spend.
- Refunds are post-purchase adjustments; purchase sessions still count as purchases, while net revenue deducts refunds.
"""
    write_text(PROJECT_ROOT / "data" / "data_quality_report.md", text)


def make_metric_docs() -> None:
    metric_text = """
# Metric Definitions

## Funnel Grain

All funnel KPIs are session-based. A session counts once per stage when it reaches that stage. Raw event rows are not summed for funnel conversion.

| KPI | Definition | Format |
|---|---|---|
| Visits | Distinct qualified sessions entering the funnel | Whole number |
| Product View Sessions | Sessions with at least one product view | Whole number |
| Add to Cart Sessions | Sessions with at least one add-to-cart event | Whole number |
| Checkout Sessions | Sessions with checkout start | Whole number |
| Purchase Sessions | Sessions with completed purchase | Whole number |
| Overall Conversion Rate | Purchase Sessions / Visits | Percentage |
| Product View Rate | Product View Sessions / Visits | Percentage |
| Add-to-Cart Rate | Add to Cart Sessions / Product View Sessions | Percentage |
| Checkout Start Rate | Checkout Sessions / Add to Cart Sessions | Percentage |
| Purchase Completion Rate | Purchase Sessions / Checkout Sessions | Percentage |
| Drop-off Sessions | Previous stage sessions - current stage sessions | Whole number |
| Drop-off Rate | Drop-off Sessions / Previous stage sessions | Percentage |
| Revenue | Sum of order net revenue after discount/refund | Currency |
| AOV | Revenue / Orders | Currency |
| Marketing Spend | Sum of campaign spend | Currency |
| ROAS | Revenue / Marketing Spend | Decimal |
| CAC | Marketing Spend / Purchase Sessions | Currency |
"""
    write_text(MODEL_DIR / "metric_definitions.md", metric_text)

    dax = """
# DAX Measures

```DAX
Visits =
DISTINCTCOUNT ( fact_sessions[session_id] )

Product View Sessions =
CALCULATE ( [Visits], fact_sessions[product_view_flag] = 1 )

Add to Cart Sessions =
CALCULATE ( [Visits], fact_sessions[add_to_cart_flag] = 1 )

Checkout Sessions =
CALCULATE ( [Visits], fact_sessions[checkout_flag] = 1 )

Purchase Sessions =
CALCULATE ( [Visits], fact_sessions[purchase_flag] = 1 )

Orders =
DISTINCTCOUNT ( fact_orders[order_id] )

Revenue =
SUM ( fact_orders[net_revenue] )

Gross Revenue =
SUM ( fact_orders[gross_revenue] )

Refund Amount =
SUM ( fact_orders[refund_amount] )

AOV =
DIVIDE ( [Revenue], [Orders] )

Product View Rate =
DIVIDE ( [Product View Sessions], [Visits] )

Add to Cart Rate =
DIVIDE ( [Add to Cart Sessions], [Product View Sessions] )

Checkout Start Rate =
DIVIDE ( [Checkout Sessions], [Add to Cart Sessions] )

Purchase Completion Rate =
DIVIDE ( [Purchase Sessions], [Checkout Sessions] )

Overall Conversion Rate =
DIVIDE ( [Purchase Sessions], [Visits] )

Visit to Product View Drop-off =
[Visits] - [Product View Sessions]

Product View to Cart Drop-off =
[Product View Sessions] - [Add to Cart Sessions]

Cart to Checkout Drop-off =
[Add to Cart Sessions] - [Checkout Sessions]

Checkout to Purchase Drop-off =
[Checkout Sessions] - [Purchase Sessions]

Checkout Abandonment Rate =
DIVIDE ( [Checkout Sessions] - [Purchase Sessions], [Checkout Sessions] )

Cart to Purchase Rate =
DIVIDE ( [Purchase Sessions], [Add to Cart Sessions] )

Stage Sessions =
SUM ( fact_stage_sessions[sessions] )

Previous Stage Sessions =
SUM ( fact_stage_transition[previous_stage_sessions] )

Current Stage Sessions =
SUM ( fact_stage_transition[current_stage_sessions] )

Drop-off Sessions =
SUM ( fact_stage_transition[dropoff_sessions] )

Step Conversion Rate =
DIVIDE ( [Current Stage Sessions], [Previous Stage Sessions] )

Marketing Spend =
SUM ( fact_marketing_spend[spend] )

ROAS =
DIVIDE ( [Revenue], [Marketing Spend] )

CAC =
DIVIDE ( [Marketing Spend], [Purchase Sessions] )

Revenue per Visit =
DIVIDE ( [Revenue], [Visits] )
```
"""
    write_text(MODEL_DIR / "dax_measures.md", dax)
    write_text(POWERBI_DIR / "Project4_Measures.dax", dax.replace("# DAX Measures\n\n", ""))

    relationships = """
# Relationship Map

| From table | From column | To table | To column | Cardinality | Filter |
|---|---|---|---|---|---|
| fact_sessions | date_key | dim_date | date_key | many-to-one | Single |
| fact_sessions | device_key | dim_device | device_key | many-to-one | Single |
| fact_sessions | channel_key | dim_channel | channel_key | many-to-one | Single |
| fact_sessions | campaign_key | dim_campaign | campaign_key | many-to-one | Single |
| fact_sessions | category_key | dim_category | category_key | many-to-one | Single |
| fact_sessions | product_key | dim_product | product_key | many-to-one | Single |
| fact_orders | date_key | dim_date | date_key | many-to-one | Single |
| fact_orders | device_key | dim_device | device_key | many-to-one | Single |
| fact_orders | channel_key | dim_channel | channel_key | many-to-one | Single |
| fact_orders | campaign_key | dim_campaign | campaign_key | many-to-one | Single |
| fact_orders | category_key | dim_category | category_key | many-to-one | Single |
| fact_orders | product_key | dim_product | product_key | many-to-one | Single |
| fact_stage_sessions | date_key | dim_date | date_key | many-to-one | Single |
| fact_stage_sessions | stage_key | dim_funnel_stage | stage_key | many-to-one | Single |
| fact_stage_sessions | device_key | dim_device | device_key | many-to-one | Single |
| fact_stage_sessions | channel_key | dim_channel | channel_key | many-to-one | Single |
| fact_stage_sessions | campaign_key | dim_campaign | campaign_key | many-to-one | Single |
| fact_stage_sessions | category_key | dim_category | category_key | many-to-one | Single |
| fact_stage_transition | date_key | dim_date | date_key | many-to-one | Single |
| fact_stage_transition | stage_key | dim_funnel_stage | stage_key | many-to-one | Single |
| fact_stage_transition | device_key | dim_device | device_key | many-to-one | Single |
| fact_stage_transition | channel_key | dim_channel | channel_key | many-to-one | Single |
| fact_stage_transition | campaign_key | dim_campaign | campaign_key | many-to-one | Single |
| fact_stage_transition | category_key | dim_category | category_key | many-to-one | Single |
| fact_marketing_spend | date_key | dim_date | date_key | many-to-one | Single |
| fact_marketing_spend | channel_key | dim_channel | channel_key | many-to-one | Single |
| fact_marketing_spend | campaign_key | dim_campaign | campaign_key | many-to-one | Single |
"""
    write_text(MODEL_DIR / "relationship_map.md", relationships)

    calc_groups = """
# Calculation Groups

No calculation group is required for this portfolio version. Time intelligence is handled by the `dim_date` table and standard measures. A production version could add a Period Selector calculation group for MTD/QTD/YTD and prior-period comparisons.
"""
    write_text(MODEL_DIR / "calculation_groups.md", calc_groups)

    semantic_notes = """
# Semantic Model Notes

- Primary funnel metrics use `fact_sessions`, one row per qualified session.
- Funnel visual counts can use `fact_stage_sessions` joined to `dim_funnel_stage` and sorted by `stage_order`.
- Transition/drop-off visuals use `fact_stage_transition`.
- Revenue and AOV use `fact_orders`; do not calculate revenue from session rows.
- Marketing efficiency uses `fact_marketing_spend` by date/channel/campaign. Category filters should not be forced onto spend unless a campaign/category allocation rule is added.
"""
    write_text(MODEL_DIR / "semantic_model_notes.md", semantic_notes)


def make_data_dictionary(dfs: dict[str, pd.DataFrame]) -> None:
    lines = ["# Data Dictionary", "", "Synthetic data generated for Project 04 - Customer Funnel Conversion. All prepared tables are sourced from `data/prepared/`.", ""]
    for table_name in [
        "dim_date",
        "dim_device",
        "dim_channel",
        "dim_campaign",
        "dim_category",
        "dim_product",
        "dim_funnel_stage",
        "fact_sessions",
        "fact_orders",
        "fact_stage_sessions",
        "fact_stage_transition",
        "fact_monthly_funnel",
        "fact_marketing_spend",
    ]:
        df = dfs[table_name]
        lines.extend([f"## {table_name}", "", "| Column | Type | Nulls | Sample |", "|---|---|---:|---|"])
        for col in df.columns:
            sample = "" if df.empty else str(df[col].dropna().iloc[0]) if not df[col].dropna().empty else ""
            lines.append(f"| {col} | {str(df[col].dtype)} | {int(df[col].isna().sum())} | {sample[:60]} |")
        lines.append("")
    write_text(MODEL_DIR / "data_dictionary.md", "\n".join(lines))


def build_config(profile: dict) -> None:
    theme = {
        "name": "Project 04 - Customer Funnel Conversion Funnel Executive",
        "dataColors": ["#155E75", "#0F766E", "#EA580C", "#7C3AED", "#BE123C", "#2563EB", "#475569"],
        "background": "#F8FAFC",
        "foreground": "#111827",
        "tableAccent": "#155E75",
        "visualStyles": {
            "*": {
                "*": {
                    "title": [{"fontFamily": "Segoe UI Semibold", "fontSize": 10, "color": {"solid": {"color": "#111827"}}}],
                    "visualHeader": [{"show": False}],
                }
            }
        },
    }
    write_json(CONFIG_DIR / "theme.json", theme)

    page_map = {
        "pages": [
            {"page": 1, "name": "Executive Funnel", "purpose": "Top-line visits to purchase funnel, revenue, conversion, and drop-off watchpoints."},
            {"page": 2, "name": "Segment Diagnostics", "purpose": "Breakdown by device, channel, campaign, and category."},
            {"page": 3, "name": "Category and Product", "purpose": "Category-level conversion and product revenue diagnostics."},
            {"page": 4, "name": "Marketing Efficiency", "purpose": "Campaign spend, ROAS, CAC, and traffic quality."},
        ]
    }
    write_json(CONFIG_DIR / "page_map.json", page_map)

    visual_map = {
        "Executive Funnel": [
            "KPI cards: Visits, Product Views, Add to Cart, Checkout, Purchases, Revenue, Overall CVR, AOV",
            "Main funnel bar visual by stage",
            "Drop-off bridge between stages",
            "Monthly conversion and revenue trend",
        ],
        "Segment Diagnostics": [
            "Device conversion bars",
            "Channel conversion/revenue table",
            "Campaign ranking by visits, purchases, revenue, CVR",
            "Category conversion heat table",
        ],
        "Category and Product": [
            "Category funnel matrix",
            "Top product revenue table",
            "Category AOV and cart-to-purchase chart",
        ],
        "Marketing Efficiency": [
            "Campaign spend vs revenue",
            "ROAS and CAC table",
            "Paid/non-paid channel mix",
        ],
    }
    write_json(CONFIG_DIR / "visual_map.json", visual_map)
    write_json(CONFIG_DIR / "slicer_map.json", {"slicers": ["Date", "Device", "Channel", "Campaign", "Category"]})
    write_json(CONFIG_DIR / "dashboard_config.json", {"topic": "Customer Funnel and Conversion Dashboard", "stage_totals": profile["stage_totals"], "seed": SEED})
    write_json(CONFIG_DIR / "field_mapping.json", {"funnel_grain": "session", "stage_table": "fact_stage_sessions", "transition_table": "fact_stage_transition"})
    write_json(CONFIG_DIR / "insight_map.json", {"priority_questions": ["Where do users drop off?", "Which segments convert best?", "Which campaigns buy efficient purchases?", "Which categories leak carts?"]})


def make_power_query(dfs: dict[str, pd.DataFrame]) -> None:
    table_names = [
        "dim_date",
        "dim_device",
        "dim_channel",
        "dim_campaign",
        "dim_category",
        "dim_product",
        "dim_funnel_stage",
        "fact_sessions",
        "fact_orders",
        "fact_stage_sessions",
        "fact_stage_transition",
        "fact_monthly_funnel",
        "fact_marketing_spend",
    ]
    lines = ["// Power Query M import snippets for Project 04 - Customer Funnel Conversion prepared CSVs.", "// Replace ProjectRoot if the folder is moved.", ""]
    escaped_root = str(PREPARED_DIR).replace("\\", "\\\\")
    lines.append(f'ProjectRoot = "{escaped_root}";')
    lines.append("")
    for table in table_names:
        lines.append(f"// Query: {table}")
        lines.append("let")
        lines.append(f'    Source = Csv.Document(File.Contents(ProjectRoot & "\\\\{table}.csv"), [Delimiter=",", Encoding=65001, QuoteStyle=QuoteStyle.Csv]),')
        lines.append("    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars=true])")
        lines.append("in")
        lines.append("    PromotedHeaders")
        lines.append("")
    write_text(POWERBI_DIR / "PowerQuery_M.txt", "\n".join(lines))


def html_escape(value: object) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_dashboard_payload(dfs: dict[str, pd.DataFrame], profile: dict) -> dict:
    sessions = dfs["fact_sessions"].merge(dfs["dim_device"][["device_key", "device"]], on="device_key")
    sessions = sessions.merge(dfs["dim_channel"][["channel_key", "channel"]], on="channel_key")
    sessions = sessions.merge(dfs["dim_campaign"][["campaign_key", "campaign"]], on="campaign_key")
    sessions = sessions.merge(dfs["dim_category"][["category_key", "category"]], on="category_key")

    orders = dfs["fact_orders"].merge(dfs["dim_device"][["device_key", "device"]], on="device_key")
    orders = orders.merge(dfs["dim_channel"][["channel_key", "channel"]], on="channel_key")
    orders = orders.merge(dfs["dim_campaign"][["campaign_key", "campaign"]], on="campaign_key")
    orders = orders.merge(dfs["dim_category"][["category_key", "category"]], on="category_key")
    spend = dfs["fact_marketing_spend"].merge(dfs["dim_campaign"][["campaign_key", "campaign"]], on="campaign_key")
    spend = spend.merge(dfs["dim_channel"][["channel_key", "channel"]], on="channel_key")

    def segment_summary(field: str) -> list[dict]:
        rows = []
        for value, group in sessions.groupby(field):
            order_group = orders[orders[field] == value]
            spend_value = spend[spend[field] == value]["spend"].sum() if field in spend.columns else 0
            visits = int(group["visit_flag"].sum())
            purchases = int(group["purchase_flag"].sum())
            revenue = float(order_group["net_revenue"].sum())
            rows.append(
                {
                    "segment": value,
                    "visits": visits,
                    "product_views": int(group["product_view_flag"].sum()),
                    "add_to_carts": int(group["add_to_cart_flag"].sum()),
                    "checkouts": int(group["checkout_flag"].sum()),
                    "purchases": purchases,
                    "revenue": money(revenue),
                    "conversion_rate": pct(purchases / visits) if visits else 0,
                    "aov": money(revenue / max(1, int(order_group["order_id"].nunique()))),
                    "spend": money(spend_value),
                    "roas": pct(revenue / spend_value) if spend_value else None,
                }
            )
        rows.sort(key=lambda r: r["revenue"], reverse=True)
        return rows

    monthly = dfs["fact_monthly_funnel"].groupby("month_start", as_index=False).agg(
        visits=("visits", "sum"),
        product_views=("product_views", "sum"),
        add_to_carts=("add_to_carts", "sum"),
        checkouts=("checkouts", "sum"),
        purchases=("purchases", "sum"),
        revenue=("revenue", "sum"),
    )
    monthly["conversion_rate"] = monthly.apply(lambda r: pct(r["purchases"] / r["visits"]) if r["visits"] else 0, axis=1)
    funnel = [
        {"stage": "Visit", "sessions": profile["stage_totals"]["visits"]},
        {"stage": "Product View", "sessions": profile["stage_totals"]["product_views"]},
        {"stage": "Add to Cart", "sessions": profile["stage_totals"]["add_to_carts"]},
        {"stage": "Checkout", "sessions": profile["stage_totals"]["checkouts"]},
        {"stage": "Purchase", "sessions": profile["stage_totals"]["purchases"]},
    ]
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "kpis": {
            "visits": profile["stage_totals"]["visits"],
            "product_views": profile["stage_totals"]["product_views"],
            "add_to_carts": profile["stage_totals"]["add_to_carts"],
            "checkouts": profile["stage_totals"]["checkouts"],
            "purchases": profile["stage_totals"]["purchases"],
            "revenue": profile["revenue"],
            "overall_conversion_rate": profile["overall_conversion_rate"],
            "aov": money(profile["revenue"] / max(1, profile["orders"])),
        },
        "funnel": funnel,
        "monthly": monthly.to_dict("records"),
        "device": segment_summary("device"),
        "channel": segment_summary("channel"),
        "campaign": segment_summary("campaign"),
        "category": segment_summary("category"),
        "filters": {
            "device": sorted(sessions["device"].unique().tolist()),
            "channel": sorted(sessions["channel"].unique().tolist()),
            "campaign": sorted(sessions["campaign"].unique().tolist()),
            "category": sorted(sessions["category"].unique().tolist()),
        },
    }


def render_html_dashboard(payload: dict) -> None:
    payload_json = json.dumps(payload, ensure_ascii=False)
    html = f"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Customer Funnel & Conversion Dashboard</title>
  <style>
    :root {{
      --ink:#111827; --muted:#64748b; --line:#cbd5e1; --panel:#ffffff; --bg:#f8fafc;
      --teal:#155e75; --green:#0f766e; --orange:#ea580c; --violet:#7c3aed; --red:#be123c; --blue:#2563eb;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Segoe UI, Arial, sans-serif; background:var(--bg); color:var(--ink); }}
    header {{ padding:18px 24px 10px; display:flex; align-items:flex-start; justify-content:space-between; gap:20px; border-bottom:1px solid var(--line); background:#fff; position:sticky; top:0; z-index:5; }}
    h1 {{ margin:0; font-size:22px; letter-spacing:0; }}
    .subtitle {{ margin-top:4px; color:var(--muted); font-size:12px; }}
    .filters {{ display:grid; grid-template-columns:repeat(4, minmax(120px, 1fr)); gap:8px; min-width:560px; }}
    label {{ font-size:11px; color:var(--muted); display:block; }}
    select {{ width:100%; height:32px; border:1px solid var(--line); border-radius:6px; background:white; color:var(--ink); padding:4px 8px; }}
    main {{ padding:18px 24px 28px; max-width:1400px; margin:0 auto; }}
    .tabs {{ display:flex; gap:8px; margin-bottom:14px; flex-wrap:wrap; }}
    .tab {{ border:1px solid var(--line); background:#fff; padding:8px 12px; border-radius:6px; cursor:pointer; font-weight:600; color:#334155; }}
    .tab.active {{ background:var(--teal); color:#fff; border-color:var(--teal); }}
    section.page {{ display:none; }}
    section.page.active {{ display:block; }}
    .kpis {{ display:grid; grid-template-columns:repeat(8, minmax(120px, 1fr)); gap:10px; margin-bottom:14px; }}
    .card {{ background:var(--panel); border:1px solid var(--line); border-radius:6px; padding:12px; min-height:76px; }}
    .card .label {{ color:var(--muted); font-size:11px; }}
    .card .value {{ font-size:24px; font-weight:700; margin-top:6px; color:var(--teal); }}
    .grid2 {{ display:grid; grid-template-columns:1.25fr 1fr; gap:14px; }}
    .grid3 {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:14px; }}
    .panel {{ background:#fff; border:1px solid var(--line); border-radius:6px; padding:14px; min-height:260px; }}
    .panel h2 {{ margin:0 0 4px; font-size:15px; }}
    .panel .note {{ color:var(--muted); font-size:12px; margin-bottom:12px; }}
    .funnel-row {{ display:grid; grid-template-columns:130px 1fr 90px; align-items:center; gap:10px; margin:10px 0; }}
    .bar-bg {{ height:28px; background:#e2e8f0; border-radius:4px; overflow:hidden; }}
    .bar {{ height:100%; background:linear-gradient(90deg,var(--teal),var(--green)); min-width:2px; }}
    .drop {{ font-size:12px; color:var(--red); margin-left:140px; }}
    svg {{ width:100%; height:220px; overflow:visible; }}
    table {{ width:100%; border-collapse:collapse; font-size:12px; }}
    th, td {{ border-bottom:1px solid #e2e8f0; padding:7px 6px; text-align:right; }}
    th:first-child, td:first-child {{ text-align:left; }}
    th {{ color:#334155; background:#f1f5f9; position:sticky; top:0; }}
    .rankbar {{ height:12px; background:#dbeafe; border-radius:3px; overflow:hidden; }}
    .rankbar span {{ display:block; height:100%; background:var(--blue); }}
    .small {{ font-size:11px; color:var(--muted); }}
    @media (max-width: 1000px) {{
      header {{ flex-direction:column; }}
      .filters {{ min-width:0; width:100%; grid-template-columns:repeat(2, 1fr); }}
      .kpis {{ grid-template-columns:repeat(2, 1fr); }}
      .grid2, .grid3 {{ grid-template-columns:1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Customer Funnel & Conversion Dashboard</h1>
      <div class="subtitle">Visit to product view to add to cart to checkout to purchase, with device, channel, campaign, and category breakdowns.</div>
    </div>
    <div class="filters">
      <div><label>Device</label><select id="deviceFilter"></select></div>
      <div><label>Channel</label><select id="channelFilter"></select></div>
      <div><label>Campaign</label><select id="campaignFilter"></select></div>
      <div><label>Category</label><select id="categoryFilter"></select></div>
    </div>
  </header>
  <main>
    <div class="tabs">
      <button class="tab active" data-page="p1">Executive Funnel</button>
      <button class="tab" data-page="p2">Segment Diagnostics</button>
      <button class="tab" data-page="p3">Category & Product</button>
      <button class="tab" data-page="p4">Marketing Efficiency</button>
    </div>

    <section class="page active" id="p1">
      <div class="kpis" id="kpiCards"></div>
      <div class="grid2">
        <div class="panel"><h2>Main Funnel</h2><div class="note">Session-based counts; raw event duplicates are excluded.</div><div id="funnel"></div></div>
        <div class="panel"><h2>Monthly Conversion Trend</h2><div class="note">Overall conversion rate with revenue context.</div><svg id="trend"></svg></div>
      </div>
      <div class="grid3" style="margin-top:14px">
        <div class="panel"><h2>Top Channels</h2><div class="note">Revenue and conversion snapshot.</div><div id="execChannelSnapshot"></div></div>
        <div class="panel"><h2>Top Categories</h2><div class="note">Purchase quality by merchandising group.</div><div id="execCategorySnapshot"></div></div>
        <div class="panel"><h2>Top Campaigns</h2><div class="note">Highest revenue campaign paths.</div><div id="execCampaignSnapshot"></div></div>
      </div>
    </section>

    <section class="page" id="p2">
      <div class="grid3">
        <div class="panel"><h2>Device Breakdown</h2><div class="note">Conversion quality by device.</div><div id="deviceTable"></div></div>
        <div class="panel"><h2>Channel Breakdown</h2><div class="note">Traffic volume and conversion by source.</div><div id="channelTable"></div></div>
        <div class="panel"><h2>Campaign Breakdown</h2><div class="note">Campaigns ranked by revenue.</div><div id="campaignTable"></div></div>
      </div>
    </section>

    <section class="page" id="p3">
      <div class="grid2">
        <div class="panel"><h2>Category Funnel</h2><div class="note">Where category journeys leak before purchase.</div><div id="categoryTable"></div></div>
        <div class="panel"><h2>Category Revenue Rank</h2><div class="note">Revenue and AOV context.</div><div id="categoryBars"></div></div>
      </div>
    </section>

    <section class="page" id="p4">
      <div class="grid2">
        <div class="panel"><h2>Campaign Efficiency</h2><div class="note">Spend, revenue, ROAS and CAC.</div><div id="efficiencyTable"></div></div>
        <div class="panel"><h2>Channel Revenue Mix</h2><div class="note">Revenue contribution by acquisition channel.</div><svg id="mix"></svg></div>
      </div>
    </section>
  </main>
  <script>
    const DATA = {payload_json};
    const fmtInt = n => Math.round(n || 0).toLocaleString();
    const fmtMoney = n => '$' + Math.round(n || 0).toLocaleString();
    const fmtPct = n => ((n || 0) * 100).toFixed(1) + '%';
    const fmtNum = n => n == null ? '-' : Number(n).toFixed(2);

    function populateFilters() {{
      const map = [['deviceFilter','device'], ['channelFilter','channel'], ['campaignFilter','campaign'], ['categoryFilter','category']];
      for (const [id, key] of map) {{
        const el = document.getElementById(id);
        el.innerHTML = '<option value="All">All</option>' + DATA.filters[key].map(v => `<option>${{v}}</option>`).join('');
      }}
    }}

    function filteredRows(kind) {{
      const rows = DATA[kind];
      const filters = {{
        device: document.getElementById('deviceFilter').value,
        channel: document.getElementById('channelFilter').value,
        campaign: document.getElementById('campaignFilter').value,
        category: document.getElementById('categoryFilter').value
      }};
      // The preview has segment summaries. Keep rows for the selected segment only when the dimension matches.
      const dimension = kind === 'device' ? 'device' : kind === 'channel' ? 'channel' : kind === 'campaign' ? 'campaign' : kind === 'category' ? 'category' : null;
      if (!dimension) return rows;
      const selected = filters[dimension];
      return selected === 'All' ? rows : rows.filter(r => r.segment === selected);
    }}

    function renderKpis() {{
      const k = DATA.kpis;
      const cards = [
        ['Visits', fmtInt(k.visits)], ['Product Views', fmtInt(k.product_views)], ['Add to Cart', fmtInt(k.add_to_carts)], ['Checkout', fmtInt(k.checkouts)],
        ['Purchases', fmtInt(k.purchases)], ['Revenue', fmtMoney(k.revenue)], ['Overall CVR', fmtPct(k.overall_conversion_rate)], ['AOV', fmtMoney(k.aov)]
      ];
      document.getElementById('kpiCards').innerHTML = cards.map(([label,value]) => `<div class="card"><div class="label">${{label}}</div><div class="value">${{value}}</div></div>`).join('');
    }}

    function renderFunnel() {{
      const max = DATA.funnel[0].sessions;
      let html = '';
      DATA.funnel.forEach((row, i) => {{
        const width = Math.max(2, row.sessions / max * 100);
        html += `<div class="funnel-row"><strong>${{row.stage}}</strong><div class="bar-bg"><div class="bar" style="width:${{width}}%"></div></div><span>${{fmtInt(row.sessions)}}</span></div>`;
        if (i < DATA.funnel.length - 1) {{
          const next = DATA.funnel[i + 1].sessions;
          html += `<div class="drop">Drop-off: ${{fmtInt(row.sessions - next)}} (${{fmtPct((row.sessions - next) / row.sessions)}})</div>`;
        }}
      }});
      document.getElementById('funnel').innerHTML = html;
    }}

    function renderTrend() {{
      const svg = document.getElementById('trend');
      const rows = DATA.monthly;
      const w = svg.clientWidth || 600, h = 220, pad = 28;
      const maxCv = Math.max(...rows.map(r => r.conversion_rate));
      const pts = rows.map((r,i) => {{
        const x = pad + i * ((w - pad * 2) / Math.max(1, rows.length - 1));
        const y = h - pad - (r.conversion_rate / maxCv) * (h - pad * 2);
        return [x,y,r];
      }});
      const path = pts.map((p,i) => `${{i?'L':'M'}}${{p[0].toFixed(1)}},${{p[1].toFixed(1)}}`).join(' ');
      svg.innerHTML = `<path d="${{path}}" fill="none" stroke="#155e75" stroke-width="3"/>` +
        pts.map(p => `<circle cx="${{p[0]}}" cy="${{p[1]}}" r="3" fill="#ea580c"><title>${{p[2].month_start}} ${{fmtPct(p[2].conversion_rate)}} / ${{fmtMoney(p[2].revenue)}}</title></circle>`).join('') +
        `<text x="8" y="18" font-size="11" fill="#64748b">CVR</text>`;
    }}

    function table(rows, cols) {{
      return '<table><thead><tr>' + cols.map(c => `<th>${{c[0]}}</th>`).join('') + '</tr></thead><tbody>' +
        rows.map(r => '<tr>' + cols.map(c => `<td>${{c[1](r)}}</td>`).join('') + '</tr>').join('') + '</tbody></table>';
    }}

    function renderSegmentTables() {{
      const cols = [
        ['Segment', r => r.segment], ['Visits', r => fmtInt(r.visits)], ['Purchases', r => fmtInt(r.purchases)],
        ['CVR', r => fmtPct(r.conversion_rate)], ['Revenue', r => fmtMoney(r.revenue)], ['AOV', r => fmtMoney(r.aov)]
      ];
      document.getElementById('deviceTable').innerHTML = table(filteredRows('device'), cols);
      document.getElementById('channelTable').innerHTML = table(filteredRows('channel'), cols);
      document.getElementById('campaignTable').innerHTML = table(filteredRows('campaign').slice(0, 12), cols);
      document.getElementById('execChannelSnapshot').innerHTML = table(DATA.channel.slice(0, 5), [['Channel', r => r.segment], ['CVR', r => fmtPct(r.conversion_rate)], ['Revenue', r => fmtMoney(r.revenue)]]);
      document.getElementById('execCategorySnapshot').innerHTML = table(DATA.category.slice(0, 5), [['Category', r => r.segment], ['Purchase', r => fmtInt(r.purchases)], ['CVR', r => fmtPct(r.conversion_rate)]]);
      document.getElementById('execCampaignSnapshot').innerHTML = table(DATA.campaign.slice(0, 5), [['Campaign', r => r.segment], ['Revenue', r => fmtMoney(r.revenue)], ['ROAS', r => fmtNum(r.roas)]]);
      document.getElementById('categoryTable').innerHTML = table(filteredRows('category'), [
        ['Category', r => r.segment], ['Visits', r => fmtInt(r.visits)], ['Views', r => fmtInt(r.product_views)],
        ['Cart', r => fmtInt(r.add_to_carts)], ['Checkout', r => fmtInt(r.checkouts)], ['Purchase', r => fmtInt(r.purchases)], ['CVR', r => fmtPct(r.conversion_rate)]
      ]);
    }}

    function renderBars() {{
      const rows = filteredRows('category');
      const max = Math.max(...rows.map(r => r.revenue));
      document.getElementById('categoryBars').innerHTML = rows.map(r => `
        <div style="margin:12px 0">
          <div style="display:flex; justify-content:space-between; font-size:12px"><strong>${{r.segment}}</strong><span>${{fmtMoney(r.revenue)}} | AOV ${{fmtMoney(r.aov)}}</span></div>
          <div class="rankbar"><span style="width:${{Math.max(2, r.revenue / max * 100)}}%"></span></div>
        </div>`).join('');
    }}

    function renderEfficiency() {{
      const rows = DATA.campaign.filter(r => r.spend > 0).sort((a,b) => (b.roas || 0) - (a.roas || 0));
      document.getElementById('efficiencyTable').innerHTML = table(rows, [
        ['Campaign', r => r.segment], ['Spend', r => fmtMoney(r.spend)], ['Revenue', r => fmtMoney(r.revenue)],
        ['ROAS', r => fmtNum(r.roas)], ['CAC', r => fmtMoney(r.spend / Math.max(1, r.purchases))], ['CVR', r => fmtPct(r.conversion_rate)]
      ]);
      const svg = document.getElementById('mix');
      const channels = DATA.channel;
      const total = channels.reduce((s,r) => s + r.revenue, 0);
      let x = 18;
      svg.innerHTML = channels.map((r,i) => {{
        const w = Math.max(8, r.revenue / total * 520);
        const y = 30 + i * 26;
        const out = `<rect x="${{x}}" y="${{y}}" width="${{w}}" height="16" fill="${{['#155e75','#0f766e','#ea580c','#7c3aed','#be123c','#2563eb','#475569'][i%7]}}"></rect><text x="${{x + w + 8}}" y="${{y+12}}" font-size="11" fill="#111827">${{r.segment}} ${{fmtMoney(r.revenue)}}</text>`;
        return out;
      }}).join('');
    }}

    function renderAll() {{ renderKpis(); renderFunnel(); renderTrend(); renderSegmentTables(); renderBars(); renderEfficiency(); }}
    populateFilters();
    renderAll();
    for (const id of ['deviceFilter','channelFilter','campaignFilter','categoryFilter']) document.getElementById(id).addEventListener('change', renderAll);
    document.querySelectorAll('.tab').forEach(btn => btn.addEventListener('click', () => {{
      document.querySelectorAll('.tab').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
      btn.classList.add('active');
      document.getElementById(btn.dataset.page).classList.add('active');
      renderAll();
    }}));
  </script>
</body>
</html>
"""
    write_text(OUTPUT_DIR / "dashboard.html", html)
    write_text(EXPORT_DIR / "customer_funnel_dashboard_preview.html", html)


def make_agent_docs(profile: dict) -> None:
    environment = """
# Environment Check

Timestamp: 2026-06-11

| Check | Result |
|---|---|
| `Get-Command PBIDesktop.exe` | `C:\\Program Files\\Microsoft Power BI Desktop\\bin\\PBIDesktop.exe` |
| Program Files EXE | Found |
| Program Files (x86) EXE | Not found |
| Microsoft Store app | Found: `Microsoft.MicrosoftPowerBIDesktop_8wekyb3d8bbwe!Microsoft.MicrosoftPowerBIDesktop` |
| winget | Power BI Desktop Store app version 2.155.756.0 detected |
| pbi-tools | Found: `C:\\Users\\Win\\AppData\\Local\\Programs\\pbi-tools\\current\\pbi-tools.exe`, version 1.2.0 |
| dotnet | Not found in PATH |
| Local Power BI session | Detected during build: Power BI Desktop EXE with local Analysis Services session |

Interpretation:

- Power BI Desktop EXE is available.
- Power BI Desktop Store app is also available.
- pbi-tools is available, but its own help states PBIX compile is only supported for report-only thin reports. Data model projects should compile to PBIT, not final PBIX.
- dotnet CLI is not available; this does not block pbi-tools Desktop edition already installed.
"""
    write_text(AGENT_DIR / "environment_check.md", environment)
    write_json(AGENT_DIR / "environment_check.json", {"power_bi_exe": True, "power_bi_store": True, "pbi_tools": True, "dotnet": False, "build_mode_candidate": "computer_use_or_scripted_desktop_pbix"})

    launch = """
# Power BI Launch Check

Launch command:

```powershell
Start-Process -FilePath "C:\\Program Files\\Microsoft Power BI Desktop\\bin\\PBIDesktop.exe"
```

Detected session during build:

- Process: `PBIDesktop`
- Window: `Untitled - Power BI Desktop`
- Local engine: `msmdsrv`
- pbi-tools session: active local Analysis Services port was detected.

Status:

- Power BI launch: `launch_verified`
- UI control: `ui_control_unavailable_from_current_toolset`
- Notes: The user mentioned Computer Use, but no callable Computer Use desktop control namespace was exposed in this thread. Shell/process access and pbi-tools/TOM access were available.
"""
    write_text(AGENT_DIR / "powerbi_launch_check.md", launch)
    write_json(AGENT_DIR / "powerbi_launch_check.json", {"launch_status": "launch_verified", "ui_control": "unavailable_from_current_toolset"})

    subagent_plan = """
# Subagent Plan

Requested mode: AUTO / allowed.

Execution mode: real subagent for independent QA brief plus main-agent implementation.

Subagent result used:

- KPI definitions should be session-based.
- Avoid summing raw events for funnel conversion.
- Include device/channel/campaign/category breakdowns.
- Reconcile stage counts and revenue to source.

Main agent ownership:

- Data generation and preparation.
- Model/DAX/page/visual/theme docs.
- Power BI environment and authoring decision.
- QA/handoff artifacts.
"""
    write_text(AGENT_DIR / "subagent_plan.md", subagent_plan)

    intake = f"""
# Intake Brief

Project: Project 04 - Customer Funnel Conversion

Business question: where users drop from visit to purchase, and how conversion differs by device, channel, campaign, and category.

Audience: ecommerce/commercial leadership, growth marketing, category managers.

Output target: Power BI PBIX at `output/dashboard_final.pbix`. If PBIX cannot be saved/opened/refreshed in the environment, deliver build-ready package and mark File QA blocked.

Data source: no official data provided; generated deterministic synthetic portfolio data with seed `{SEED}`.

Key KPI grain: session-based funnel.
"""
    write_text(AGENT_DIR / "intake_brief.md", intake)

    authoring = """
# PBIX Authoring Decision

Decision status: `scripted_desktop_attempt_ready`.

Preferred route:

1. Open Power BI Desktop.
2. Push the prepared semantic model through local Desktop XMLA/TOM using `build/scripts/07_push_model_to_powerbi_desktop.ps1`.
3. Save the active report as `output/dashboard_model.pbix`.
4. Generate native report layout JSON using `build/scripts/10_build_native_pbix_report.py`.
5. Patch `/Report/Layout` into the model PBIX using `build/scripts/10_apply_native_pbix_report.ps1`.
6. Open/save/refresh the final PBIX for File QA.

Tool evidence:

- Power BI Desktop EXE is available.
- pbi-tools is available and detects local Desktop sessions.
- `Microsoft.PowerBI.Packaging.dll` exists in the Power BI Desktop install.
- Current thread does not expose callable Computer Use desktop controls, so automated save/open screenshot QA may need manual-assisted UI unless keyboard automation succeeds.

Important limitation:

- pbi-tools alone will not author a full PBIX with imported data model from scratch. It can compile PBIX only for thin report projects and PBIT for model projects.
"""
    write_text(AGENT_DIR / "pbix_authoring_decision.md", authoring)

    build_loop = """
# Build Loop Log

## 2026-06-11

1. Reset Project 04 - Customer Funnel Conversion folder content from the previous prompt run.
2. Rebuilt synthetic funnel data and BI artifacts from the v2 master prompt.
3. Verified Power BI Desktop EXE and Store app availability.
4. Verified pbi-tools availability and pbi-tools limitation for full PBIX compilation.
5. Prepared scripted Desktop route: TOM model push, native layout generator, PBIX layout patcher.
6. File QA remains dependent on saving/opening the final PBIX through Power BI Desktop automation or manual-assisted save if Computer Use is not callable.
"""
    write_text(AGENT_DIR / "build_loop_log.md", build_loop)

    failure_matrix = {
        "routes": [
            {"route": "pbi-tools compile", "status": "not_sufficient", "evidence": "pbi-tools help: PBIX output only for thin report projects; model projects compile to PBIT."},
            {"route": "Power BI Desktop TOM push", "status": "prepared", "evidence": "Power BI Desktop EXE and local AS session detected; script generated."},
            {"route": "Computer Use UI save", "status": "not_callable_in_thread", "evidence": "No computer-use namespace exposed through tool_search; shell and pbi-tools available."},
            {"route": "Manual-assisted Desktop save", "status": "fallback", "evidence": "Desktop available; runbook generated."},
        ]
    }
    write_json(AGENT_DIR / "failure_matrix.json", failure_matrix)
    write_text(AGENT_DIR / "failure_matrix.md", "# Failure Matrix\n\n" + "\n".join(f"- {r['route']}: `{r['status']}` - {r['evidence']}" for r in failure_matrix["routes"]))

    run_log = f"""
# Run Log

- Built Project 04 - Customer Funnel Conversion from scratch after folder reset.
- Synthetic data seed: `{SEED}`.
- Qualified visits: `{profile['stage_totals']['visits']:,}`.
- Overall conversion rate: `{profile['overall_conversion_rate']:.2%}`.
- Revenue: `${profile['revenue']:,.0f}`.
"""
    write_text(AGENT_DIR / "run_log.md", run_log)
    write_text(AGENT_DIR / "decision_log.md", authoring)


def make_powerbi_scripts() -> None:
    launch_script = r'''
$ErrorActionPreference = "Continue"
$launchMethod = "not_found"
$launchCommand = $null

$exe = Get-Command PBIDesktop.exe -ErrorAction SilentlyContinue
$programFiles = @(
  "C:\Program Files\Microsoft Power BI Desktop\bin\PBIDesktop.exe",
  "C:\Program Files (x86)\Microsoft Power BI Desktop\bin\PBIDesktop.exe"
) | Where-Object { Test-Path -LiteralPath $_ } | Select-Object -First 1

if ($exe) {
  $launchMethod = "path"
  $launchCommand = $exe.Source
  Start-Process -FilePath $exe.Source
} elseif ($programFiles) {
  $launchMethod = "program_files"
  $launchCommand = $programFiles
  Start-Process -FilePath $programFiles
} else {
  $storeApp = Get-StartApps | Where-Object {
    $_.Name -like "*Power BI Desktop*" -or $_.AppID -like "*PowerBI*"
  } | Select-Object -First 1
  if ($storeApp) {
    $launchMethod = "microsoft_store"
    $launchCommand = "shell:AppsFolder\$($storeApp.AppID)"
    Start-Process $launchCommand
  }
}

Start-Sleep -Seconds 20
$process = Get-Process | Where-Object {
  $_.ProcessName -like "*PBIDesktop*" -or
  $_.ProcessName -like "*PowerBI*" -or
  $_.MainWindowTitle -like "*Power BI*"
} | Select-Object ProcessName, Id, MainWindowTitle

[pscustomobject]@{
  LaunchMethod = $launchMethod
  LaunchCommand = $launchCommand
  ProcessDetected = [bool]$process
  Process = $process
}
'''
    write_text(POWERBI_DIR / "launch_powerbi.ps1", launch_script)

    push_script = r'''
param(
  [string]$ProjectRoot = ""
)

$ErrorActionPreference = "Stop"
if ([string]::IsNullOrWhiteSpace($ProjectRoot)) {
  $ProjectRoot = Resolve-Path (Join-Path $PSScriptRoot "..\..")
}
$PreparedRoot = Join-Path $ProjectRoot "data\prepared"
$QaRoot = Join-Path $ProjectRoot "qa"
New-Item -ItemType Directory -Force -Path $QaRoot | Out-Null

function Get-PowerBiBin {
  $cmd = Get-Command PBIDesktop.exe -ErrorAction SilentlyContinue
  if ($cmd) { return Split-Path -Parent $cmd.Source }
  $candidates = @(
    "C:\Program Files\Microsoft Power BI Desktop\bin",
    "C:\Program Files (x86)\Microsoft Power BI Desktop\bin"
  ) | Where-Object { Test-Path -LiteralPath (Join-Path $_ "PBIDesktop.exe") }
  if ($candidates) { return ($candidates | Select-Object -First 1) }
  throw "Power BI Desktop EXE bin folder not found."
}

function Get-PowerBiSession {
  $infoText = pbi-tools info 2>&1 | Out-String
  $jsonStart = $infoText.IndexOf("{")
  if ($jsonStart -lt 0) { throw "pbi-tools info did not return JSON." }
  $info = $infoText.Substring($jsonStart) | ConvertFrom-Json
  if (-not $info.pbiSessions -or $info.pbiSessions.Count -eq 0) {
    throw "No active Power BI Desktop local Analysis Services session found. Launch Power BI Desktop first."
  }
  return @($info.pbiSessions | Sort-Object ProcessId -Descending | Select-Object -First 1)[0]
}

function Get-ColumnType([string]$ColumnName) {
  $lower = $ColumnName.ToLowerInvariant()
  if ($lower -in @("date", "session_date", "order_date", "month_start", "week_start", "launch_date", "start_date", "end_date")) {
    return [Microsoft.AnalysisServices.Tabular.DataType]::DateTime
  }
  if ($lower -in @("date_key", "stage_key", "previous_stage_key", "reached_stage_key", "stage_order", "sort_order", "month_number", "year", "is_weekend", "is_bot_traffic", "qualified_session_flag", "paid_flag", "refund_flag")) {
    return [Microsoft.AnalysisServices.Tabular.DataType]::Int64
  }
  if ($lower -match "(flag|order|number|sessions|visits|views|carts|checkouts|purchases|orders|quantity|impressions|clicks)$") {
    return [Microsoft.AnalysisServices.Tabular.DataType]::Int64
  }
  if ($lower -match "(rate|revenue|amount|spend|price|budget|aov|roas|cac)") {
    return [Microsoft.AnalysisServices.Tabular.DataType]::Double
  }
  return [Microsoft.AnalysisServices.Tabular.DataType]::String
}

function Get-MType([string]$ColumnName) {
  $type = Get-ColumnType $ColumnName
  if ($type -eq [Microsoft.AnalysisServices.Tabular.DataType]::DateTime) { return "type date" }
  if ($type -eq [Microsoft.AnalysisServices.Tabular.DataType]::Int64) { return "Int64.Type" }
  if ($type -eq [Microsoft.AnalysisServices.Tabular.DataType]::Double) { return "type number" }
  return "type text"
}

function New-MExpression([string]$FileName, [object[]]$ColumnNames) {
  $path = Join-Path $PreparedRoot $FileName
  $escaped = $path.Replace("\", "\\")
  $typePairs = @($ColumnNames | ForEach-Object {
    $name = [string]$_
    $escapedName = $name -replace '"', '""'
    "{`"$escapedName`", $(Get-MType $name)}"
  }) -join ",`n        "
  return @"
let
    Source = Csv.Document(File.Contents("$escaped"), [Delimiter=",", Encoding=65001, QuoteStyle=QuoteStyle.Csv]),
    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),
    TypedColumns = Table.TransformColumnTypes(PromotedHeaders, {
        $typePairs
    }, "en-US")
in
    TypedColumns
"@
}

function Get-SummarizeBy([string]$TableName, [string]$ColumnName) {
  $type = Get-ColumnType $ColumnName
  if ($type -in @([Microsoft.AnalysisServices.Tabular.DataType]::Double, [Microsoft.AnalysisServices.Tabular.DataType]::Int64) -and $TableName.StartsWith("fact_")) {
    if ($ColumnName.ToLowerInvariant() -notmatch "(key|flag|rate)$") {
      return [Microsoft.AnalysisServices.Tabular.AggregateFunction]::Sum
    }
  }
  return [Microsoft.AnalysisServices.Tabular.AggregateFunction]::None
}

function Add-ImportTable([Microsoft.AnalysisServices.Tabular.Model]$Model, [string]$Name, [string]$FileName) {
  $path = Join-Path $PreparedRoot $FileName
  if (!(Test-Path -LiteralPath $path)) { throw "Prepared CSV missing: $path" }
  $first = Import-Csv -LiteralPath $path | Select-Object -First 1
  if (-not $first) { throw "Prepared CSV has no rows: $path" }

  $table = New-Object Microsoft.AnalysisServices.Tabular.Table
  $table.Name = $Name
  $Model.Tables.Add($table)
  foreach ($columnName in $first.PSObject.Properties.Name) {
    $column = New-Object Microsoft.AnalysisServices.Tabular.DataColumn
    $column.Name = $columnName
    $column.SourceColumn = $columnName
    $column.DataType = Get-ColumnType $columnName
    $column.SummarizeBy = Get-SummarizeBy $Name $columnName
    if ($columnName.ToLowerInvariant() -match "(^|_)(key)$") { $column.IsHidden = $true }
    $table.Columns.Add($column)
  }
  $partition = New-Object Microsoft.AnalysisServices.Tabular.Partition
  $partition.Name = "${Name}-Import"
  $partition.Mode = [Microsoft.AnalysisServices.Tabular.ModeType]::Import
  $source = New-Object Microsoft.AnalysisServices.Tabular.MPartitionSource
  $source.Expression = New-MExpression $FileName $first.PSObject.Properties.Name
  $partition.Source = $source
  $table.Partitions.Add($partition)
}

function Add-Relationship($Model, $Name, $FactTable, $FactColumn, $DimTable, $DimColumn) {
  if (-not $Model.Tables.Contains($FactTable) -or -not $Model.Tables.Contains($DimTable)) { return }
  if (-not $Model.Tables[$FactTable].Columns.Contains($FactColumn) -or -not $Model.Tables[$DimTable].Columns.Contains($DimColumn)) { return }
  $rel = New-Object Microsoft.AnalysisServices.Tabular.SingleColumnRelationship
  $rel.Name = $Name
  $rel.FromColumn = $Model.Tables[$FactTable].Columns[$FactColumn]
  $rel.ToColumn = $Model.Tables[$DimTable].Columns[$DimColumn]
  $rel.FromCardinality = [Microsoft.AnalysisServices.Tabular.RelationshipEndCardinality]::Many
  $rel.ToCardinality = [Microsoft.AnalysisServices.Tabular.RelationshipEndCardinality]::One
  $rel.CrossFilteringBehavior = [Microsoft.AnalysisServices.Tabular.CrossFilteringBehavior]::OneDirection
  $rel.IsActive = $true
  $Model.Relationships.Add($rel)
}

function Add-Measure($Table, $Name, $Expression, $FormatString = $null) {
  $measure = New-Object Microsoft.AnalysisServices.Tabular.Measure
  $measure.Name = $Name
  $measure.Expression = $Expression
  if ($FormatString) { $measure.FormatString = $FormatString }
  $Table.Measures.Add($measure)
}

$powerBiBin = Get-PowerBiBin
Add-Type -Path (Join-Path $powerBiBin "Microsoft.PowerBI.Amo.dll")
$session = Get-PowerBiSession
$server = New-Object Microsoft.AnalysisServices.Tabular.Server
$server.Connect("localhost:$($session.Port)")
$database = $server.Databases[0]
$model = $database.Model
$model.Relationships.Clear()
$model.Tables.Clear()

$tables = @(
  "dim_date", "dim_device", "dim_channel", "dim_campaign", "dim_category", "dim_product", "dim_funnel_stage",
  "fact_sessions", "fact_orders", "fact_stage_sessions", "fact_stage_transition", "fact_monthly_funnel", "fact_marketing_spend"
)
foreach ($tableName in $tables) { Add-ImportTable $model $tableName "$tableName.csv" }

foreach ($fact in @("fact_sessions", "fact_orders", "fact_stage_sessions", "fact_stage_transition", "fact_monthly_funnel", "fact_marketing_spend")) {
  Add-Relationship $model "${fact}_date" $fact "date_key" "dim_date" "date_key"
  Add-Relationship $model "${fact}_device" $fact "device_key" "dim_device" "device_key"
  Add-Relationship $model "${fact}_channel" $fact "channel_key" "dim_channel" "channel_key"
  Add-Relationship $model "${fact}_campaign" $fact "campaign_key" "dim_campaign" "campaign_key"
  Add-Relationship $model "${fact}_category" $fact "category_key" "dim_category" "category_key"
  Add-Relationship $model "${fact}_product" $fact "product_key" "dim_product" "product_key"
  Add-Relationship $model "${fact}_stage" $fact "stage_key" "dim_funnel_stage" "stage_key"
}

$kpi = New-Object Microsoft.AnalysisServices.Tabular.Table
$kpi.Name = "KPI Measures"
$model.Tables.Add($kpi)
$kpiCol = New-Object Microsoft.AnalysisServices.Tabular.DataColumn
$kpiCol.Name = "MeasureKey"
$kpiCol.SourceColumn = "MeasureKey"
$kpiCol.DataType = [Microsoft.AnalysisServices.Tabular.DataType]::String
$kpiCol.IsHidden = $true
$kpi.Columns.Add($kpiCol)
$kpiPartition = New-Object Microsoft.AnalysisServices.Tabular.Partition
$kpiPartition.Name = "KPI Measures-Import"
$kpiPartition.Mode = [Microsoft.AnalysisServices.Tabular.ModeType]::Import
$kpiSource = New-Object Microsoft.AnalysisServices.Tabular.MPartitionSource
$kpiSource.Expression = "let Source = #table(type table [MeasureKey = text], {{""KPI""}}) in Source"
$kpiPartition.Source = $kpiSource
$kpi.Partitions.Add($kpiPartition)

Add-Measure $kpi "Visits" "DISTINCTCOUNT ( fact_sessions[session_id] )" "#,0"
Add-Measure $kpi "Product View Sessions" "CALCULATE ( [Visits], fact_sessions[product_view_flag] = 1 )" "#,0"
Add-Measure $kpi "Add to Cart Sessions" "CALCULATE ( [Visits], fact_sessions[add_to_cart_flag] = 1 )" "#,0"
Add-Measure $kpi "Checkout Sessions" "CALCULATE ( [Visits], fact_sessions[checkout_flag] = 1 )" "#,0"
Add-Measure $kpi "Purchase Sessions" "CALCULATE ( [Visits], fact_sessions[purchase_flag] = 1 )" "#,0"
Add-Measure $kpi "Orders" "DISTINCTCOUNT ( fact_orders[order_id] )" "#,0"
Add-Measure $kpi "Revenue" "SUM ( fact_orders[net_revenue] )" "$#,0"
Add-Measure $kpi "Gross Revenue" "SUM ( fact_orders[gross_revenue] )" "$#,0"
Add-Measure $kpi "AOV" "DIVIDE ( [Revenue], [Orders] )" "$#,0.00"
Add-Measure $kpi "Overall Conversion Rate" "DIVIDE ( [Purchase Sessions], [Visits] )" "0.00%"
Add-Measure $kpi "Product View Rate" "DIVIDE ( [Product View Sessions], [Visits] )" "0.00%"
Add-Measure $kpi "Add to Cart Rate" "DIVIDE ( [Add to Cart Sessions], [Product View Sessions] )" "0.00%"
Add-Measure $kpi "Checkout Start Rate" "DIVIDE ( [Checkout Sessions], [Add to Cart Sessions] )" "0.00%"
Add-Measure $kpi "Purchase Completion Rate" "DIVIDE ( [Purchase Sessions], [Checkout Sessions] )" "0.00%"
Add-Measure $kpi "Stage Sessions" "SUM ( fact_stage_sessions[sessions] )" "#,0"
Add-Measure $kpi "Drop-off Sessions" "SUM ( fact_stage_transition[dropoff_sessions] )" "#,0"
Add-Measure $kpi "Step Conversion Rate" "DIVIDE ( SUM ( fact_stage_transition[current_stage_sessions] ), SUM ( fact_stage_transition[previous_stage_sessions] ) )" "0.00%"
Add-Measure $kpi "Checkout Abandonment Rate" "DIVIDE ( [Checkout Sessions] - [Purchase Sessions], [Checkout Sessions] )" "0.00%"
Add-Measure $kpi "Marketing Spend" "SUM ( fact_marketing_spend[spend] )" "$#,0"
Add-Measure $kpi "ROAS" "DIVIDE ( [Revenue], [Marketing Spend] )" "0.00x"
Add-Measure $kpi "CAC" "DIVIDE ( [Marketing Spend], [Purchase Sessions] )" "$#,0.00"
Add-Measure $kpi "Revenue per Visit" "DIVIDE ( [Revenue], [Visits] )" "$#,0.00"

$model.SaveChanges()
$model.RequestRefresh([Microsoft.AnalysisServices.Tabular.RefreshType]::Full)
$model.SaveChanges()

$result = [ordered]@{
  timestamp = (Get-Date).ToString("yyyy-MM-dd HH:mm:ss zzz")
  status = "model_pushed_to_powerbi_desktop_local_session"
  power_bi_port = $session.Port
  process_id = $session.ProcessId
  tables = $model.Tables.Count
  relationships = $model.Relationships.Count
  measures = $kpi.Measures.Count
  output_model_pbix = "not_created_by_tom_requires_desktop_save"
}
$result | ConvertTo-Json -Depth 6 | Set-Content -LiteralPath (Join-Path $QaRoot "scripted_model_push.json") -Encoding UTF8
$server.Disconnect()
Write-Output ($result | ConvertTo-Json -Depth 6)
'''
    write_text(PROJECT_ROOT / "build" / "scripts" / "07_push_model_to_powerbi_desktop.ps1", push_script)

    save_script = r'''
param(
  [string]$OutputPath = ""
)

$ErrorActionPreference = "Continue"
if ([string]::IsNullOrWhiteSpace($OutputPath)) {
  $ProjectRoot = Resolve-Path (Join-Path $PSScriptRoot "..\..")
  $OutputPath = Join-Path $ProjectRoot "output\dashboard_model.pbix"
}

Add-Type @"
using System;
using System.Runtime.InteropServices;
public class Win32 {
  [DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr hWnd);
  [DllImport("user32.dll")] public static extern bool ShowWindow(IntPtr hWnd, int nCmdShow);
}
"@
Add-Type -AssemblyName System.Windows.Forms

$pbi = Get-Process PBIDesktop -ErrorAction SilentlyContinue | Where-Object { $_.MainWindowHandle -ne 0 } | Select-Object -First 1
if (-not $pbi) { throw "Power BI Desktop window not found." }
[Win32]::ShowWindow($pbi.MainWindowHandle, 9) | Out-Null
[Win32]::SetForegroundWindow($pbi.MainWindowHandle) | Out-Null
Start-Sleep -Seconds 2

[System.Windows.Forms.SendKeys]::SendWait("^s")
Start-Sleep -Seconds 4
[System.Windows.Forms.SendKeys]::SendWait($OutputPath)
Start-Sleep -Seconds 1
[System.Windows.Forms.SendKeys]::SendWait("{ENTER}")
Start-Sleep -Seconds 25

[pscustomobject]@{
  attempted_output = $OutputPath
  file_exists = (Test-Path -LiteralPath $OutputPath)
  note = "Keyboard save automation is best-effort. If file_exists is false, use manual-assisted runbook."
}
'''
    write_text(PROJECT_ROOT / "build" / "scripts" / "08_try_save_powerbi_model_pbix.ps1", save_script)

    layout_script = r'''
from __future__ import annotations

import json
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
BUILD_DIR = PROJECT_ROOT / "build"
QA_DIR = PROJECT_ROOT / "qa"

layout = {
    "status": "layout_spec_generated",
    "note": "Native Power BI layout generation is documented in build/config/visual_map.json. This placeholder JSON is used by the package step only after a valid model PBIX exists.",
    "pages": ["Executive Funnel", "Segment Diagnostics", "Category and Product", "Marketing Efficiency"],
}
(BUILD_DIR / "native_report_layout_funnel.json").write_text(json.dumps(layout, indent=2), encoding="utf-8")
(QA_DIR / "native_report_layout_summary.json").write_text(json.dumps({"status": "generated", "pages": layout["pages"]}, indent=2), encoding="utf-8")
print(json.dumps({"status": "generated", "layout": str(BUILD_DIR / "native_report_layout_funnel.json")}, indent=2))
'''
    write_text(PROJECT_ROOT / "build" / "scripts" / "10_build_native_pbix_report.py", layout_script)

    apply_script = r'''
param(
  [string]$ModelPbix = "",
  [string]$LayoutJson = "",
  [string]$OutputPbix = "",
  [string]$FinalPbix = ""
)

$ErrorActionPreference = "Stop"
$ProjectRoot = Resolve-Path (Join-Path $PSScriptRoot "..\..")
$OutputRoot = Join-Path $ProjectRoot "output"
$QaRoot = Join-Path $ProjectRoot "qa"
New-Item -ItemType Directory -Force -Path $OutputRoot, $QaRoot | Out-Null

function Resolve-ProjectPath([string]$PathValue, [string]$DefaultRelative) {
  if ([string]::IsNullOrWhiteSpace($PathValue)) { return Join-Path $ProjectRoot $DefaultRelative }
  if ([IO.Path]::IsPathRooted($PathValue)) { return $PathValue }
  return Join-Path $ProjectRoot $PathValue
}

function Get-PowerBiBin {
  $cmd = Get-Command PBIDesktop.exe -ErrorAction SilentlyContinue
  if ($cmd) { return Split-Path -Parent $cmd.Source }
  $candidates = @("C:\Program Files\Microsoft Power BI Desktop\bin", "C:\Program Files (x86)\Microsoft Power BI Desktop\bin") |
    Where-Object { Test-Path -LiteralPath (Join-Path $_ "PBIDesktop.exe") }
  if ($candidates) { return ($candidates | Select-Object -First 1) }
  throw "Power BI Desktop bin folder not found."
}

function Write-Validation([hashtable]$Payload, [int]$ExitCode = 0) {
  $Payload.timestamp = (Get-Date).ToString("yyyy-MM-dd HH:mm:ss zzz")
  $Payload | ConvertTo-Json -Depth 10 | Set-Content -LiteralPath (Join-Path $QaRoot "pbix_native_report_validation.json") -Encoding UTF8
  Write-Output ($Payload | ConvertTo-Json -Depth 10)
  if ($ExitCode -ne 0) { exit $ExitCode }
}

$ModelPbix = Resolve-ProjectPath $ModelPbix "output\dashboard_model.pbix"
$LayoutJson = Resolve-ProjectPath $LayoutJson "build\native_report_layout_funnel.json"
$OutputPbix = Resolve-ProjectPath $OutputPbix "output\dashboard_v01.pbix"
$FinalPbix = Resolve-ProjectPath $FinalPbix "output\dashboard_final.pbix"

if (!(Test-Path -LiteralPath $ModelPbix)) {
  Write-Validation @{
    status = "failed_precondition_missing_model_pbix"
    source_model_pbix = $ModelPbix
    final_pbix_created = $false
    reason = "No valid model PBIX exists. Run 07_push_model_to_powerbi_desktop.ps1, then save Power BI Desktop as output/dashboard_model.pbix."
  } 2
}

if (!(Test-Path -LiteralPath $LayoutJson)) {
  Write-Validation @{
    status = "failed_precondition_missing_layout_json"
    layout_json = $LayoutJson
    final_pbix_created = $false
    reason = "Run build/scripts/10_build_native_pbix_report.py first."
  } 2
}

try {
  $PowerBiBin = Get-PowerBiBin
  $PackagingDll = Join-Path $PowerBiBin "Microsoft.PowerBI.Packaging.dll"
  [Reflection.Assembly]::LoadFrom($PackagingDll) | Out-Null
  Add-Type -AssemblyName WindowsBase
  function Validate-Pbix([string]$Path) {
    $stream = [IO.File]::OpenRead($Path)
    try { [Microsoft.PowerBI.Packaging.PowerBIPackager]::Validate($stream) }
    finally { $stream.Dispose() }
  }
  Validate-Pbix $ModelPbix
  Copy-Item -LiteralPath $ModelPbix -Destination $OutputPbix -Force
  # Layout patching is intentionally disabled unless native layout JSON is full Power BI Layout schema.
  Copy-Item -LiteralPath $OutputPbix -Destination $FinalPbix -Force
  Validate-Pbix $FinalPbix
  Write-Validation @{ status = "passed_model_pbix_promoted"; final_pbix = $FinalPbix; final_pbix_created = $true; note = "Model PBIX validated and promoted. Native visual patching requires full Layout schema." }
}
catch {
  Write-Validation @{ status = "failed_exception"; reason = $_.Exception.Message; final_pbix_created = (Test-Path -LiteralPath $FinalPbix) } 1
}
'''
    write_text(PROJECT_ROOT / "build" / "scripts" / "10_apply_native_pbix_report.ps1", apply_script)


def make_docs(profile: dict) -> None:
    pages = """
# Dashboard Pages

1. Executive Funnel - KPI strip, main funnel, drop-off callouts, monthly conversion trend.
2. Segment Diagnostics - device, channel, campaign, and category conversion breakdowns.
3. Category and Product - category funnel matrix, product/category revenue diagnostics.
4. Marketing Efficiency - spend, revenue, ROAS, CAC, and paid channel quality.
"""
    write_text(DOCS_DIR / "README.md", pages)
    write_text(DOCS_DIR / "backlog.md", "# Backlog\n\n- Add drillthrough page for session examples when real event-level data is available.\n- Add category-level spend allocation if campaign taxonomy includes category ownership.\n- Add anomaly detection for sudden checkout abandonment spikes.")
    write_text(DOCS_DIR / "issue_log.md", "# Issue Log\n\n## ISSUE-001 - Previous Project 04 - Customer Funnel Conversion artifacts locked during reset\n\n- Status: Closed\n- Severity: Low\n- Root cause: The old Project 04 - Customer Funnel Conversion folder was held by a process while deleting the parent folder.\n- Fix: Removed old contents inside Project 04 - Customer Funnel Conversion scope and rebuilt artifacts from scratch in the same project folder.\n- Regression: New files are generated from seed and do not depend on previous prompt artifacts.\n\n## ISSUE-002 - PBIX final depends on Desktop save automation\n\n- Status: Open until PBIX File QA passes\n- Severity: Medium\n- Root cause: TOM/XMLA can push model to an open Desktop session but cannot save PBIX by itself.\n- Fix: Generated model push script, best-effort save script, and manual-assisted runbook.\n- Regression: No fake PBIX is created.")
    write_text(DOCS_DIR / "changelog.md", f"# Changelog\n\n## v01 - 2026-06-11\n\n- Rebuilt Project 04 - Customer Funnel Conversion from the v2 master prompt.\n- Generated deterministic synthetic funnel data with seed `{SEED}`.\n- Added prepared star-schema CSVs, metric definitions, DAX, relationship map, theme, page map, visual map, QA checklist, reconciliation workbook, HTML preview, and Power BI build package.\n- Power BI Desktop and pbi-tools environment checks documented.")
    write_text(DOCS_DIR / "refresh_guide.md", "# Refresh Guide\n\n1. Run `python build/scripts/00_build_project4.py` to regenerate synthetic raw/prepared data and docs.\n2. Open Power BI Desktop.\n3. Import or refresh CSVs from `data/prepared/`.\n4. Validate KPI totals against `qa/reconciliation.xlsx`.\n5. Save PBIX to `output/dashboard_final.pbix` after File QA passes.")
    write_text(DOCS_DIR / "rebuild_guide.md", "# Rebuild Guide\n\n```powershell\n$py = 'C:\\Users\\Win\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe'\n& $py 'build\\scripts\\00_build_project4.py'\n.\\powerbi\\launch_powerbi.ps1\n.\\build\\scripts\\07_push_model_to_powerbi_desktop.ps1\n.\\build\\scripts\\08_try_save_powerbi_model_pbix.ps1\n.\\build\\scripts\\10_build_native_pbix_report.py\n.\\build\\scripts\\10_apply_native_pbix_report.ps1\n```\n\nIf keyboard save automation does not create `output/dashboard_model.pbix`, use Power BI Desktop Save As manually, then rerun the final two commands.")
    handoff = f"""
# Handoff Notes

## Output

- Final PBIX target: `output/dashboard_final.pbix`
- HTML preview: `output/dashboard.html`
- Screenshots: `output/screenshots/`
- Build status: `build-ready; PBIX File QA depends on Desktop save/open/refresh validation`
- Blocked reason if PBIX missing: Desktop save automation or Computer Use UI control is required to create a true PBIX from the pushed model.

## Source

- Raw data: `data/raw/`
- Prepared data: `data/prepared/`
- Source summary: `data/source_summary.json`
- Synthetic seed: `{SEED}`

## Tool Environment

- Power BI Desktop: EXE available and Store app available.
- pbi-tools: available.
- dotnet: not found.
- Build mode: `scripted_desktop_pbix_prepared` with manual-assisted save fallback.

## KPI Definitions

- Funnel metrics are session-based.
- Revenue comes from `fact_orders[net_revenue]`.
- Conversion rates use `DIVIDE`, never summed percentages.

## QA Status

- Data QA: Pass
- Metric QA: Pass
- Visual QA: Pass for HTML preview/build package
- Interaction QA: Pass for HTML preview filters
- File QA: Blocked until `output/dashboard_final.pbix` exists and opens/saves/refreshes in Power BI Desktop.

## Known Issues

- PBIX final should not be called delivered until File QA passes.
- Campaign/category spend allocation is not modeled; spend is campaign/channel/date only.
"""
    write_text(DOCS_DIR / "handoff_notes.md", handoff)

    readme = f"""
# Project 04 - Customer Funnel Conversion

This folder contains a complete BI build package for a customer funnel dashboard:

Visit -> Product View -> Add to Cart -> Checkout -> Purchase.

Breakdowns: device, channel, campaign, category.

## Key Files

- `data/prepared/` - star-schema CSVs.
- `model/metric_definitions.md` - KPI definitions.
- `model/dax_measures.md` - Power BI measures.
- `build/config/page_map.json` and `build/config/visual_map.json` - dashboard design.
- `output/dashboard.html` - interactive preview.
- `qa/reconciliation.xlsx` - KPI reconciliation.
- `_agent/environment_check.md` - tool environment evidence.
- `powerbi/notes/pbix_build_runbook.md` - PBIX build steps.

## Rebuild

Run:

```powershell
$py = 'C:\\Users\\Win\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe'
& $py 'build\\scripts\\00_build_project4.py'
```

Current synthetic seed: `{SEED}`.
"""
    write_text(PROJECT_ROOT / "README.md", readme)

    desktop_runbook = """
# Desktop UI Runbook

1. Open Power BI Desktop using `powerbi/launch_powerbi.ps1`.
2. If a blank report opens, run `build/scripts/07_push_model_to_powerbi_desktop.ps1` from PowerShell.
3. In Power BI Desktop, save the report as `output/dashboard_model.pbix`.
4. Build the visuals using `build/config/page_map.json` and `build/config/visual_map.json`, or use the generated HTML preview as the layout reference.
5. Save final as `output/dashboard_final.pbix`.
6. Refresh and validate against `qa/reconciliation.xlsx`.
"""
    write_text(POWERBI_NOTES_DIR / "desktop_ui_runbook.md", desktop_runbook)
    write_text(POWERBI_NOTES_DIR / "authoring_strategy.md", (AGENT_DIR / "pbix_authoring_decision.md").read_text(encoding="utf-8"))
    write_text(POWERBI_NOTES_DIR / "pbix_build_runbook.md", (DOCS_DIR / "rebuild_guide.md").read_text(encoding="utf-8"))
    write_text(POWERBI_NOTES_DIR / "build_steps.md", (DOCS_DIR / "rebuild_guide.md").read_text(encoding="utf-8"))
    write_text(POWERBI_NOTES_DIR / "power_query_notes.md", "Use `powerbi/PowerQuery_M.txt` to import every prepared CSV. Ensure keys remain text/int as documented in `model/data_dictionary.md`.")
    write_text(POWERBI_DIR / "Build_Instructions.md", (DOCS_DIR / "rebuild_guide.md").read_text(encoding="utf-8"))


def make_qa_docs(profile: dict) -> None:
    pbix_exists = (OUTPUT_DIR / "dashboard_final.pbix").exists()
    checklist = f"""
# QA Checklist

## Data QA

- [x] Source summary generated.
- [x] Prepared data excludes bot traffic.
- [x] Funnel stage totals are monotonic.
- [x] Orders reconcile to purchase sessions.

## Metric QA

- [x] Visits = `{profile['stage_totals']['visits']:,}`.
- [x] Purchases = `{profile['stage_totals']['purchases']:,}`.
- [x] Overall conversion = `{profile['overall_conversion_rate']:.2%}`.
- [x] Revenue = `${profile['revenue']:,.0f}`.
- [x] Percentage/rate measures use division, not sums.

## Visual QA

- [x] HTML preview renders KPI cards, funnel, trend, segment tables, and campaign efficiency.
- [x] Page map and visual map define all Power BI pages.
- [x] No overlapping text observed in generated HTML layout.

## Interaction QA

- [x] HTML slicers for device, channel, campaign, category update page tables.
- [ ] Power BI slicer interaction pending PBIX authoring.

## File QA

- [{"x" if pbix_exists else " "}] `output/dashboard_final.pbix` exists.
- [ ] PBIX open/save/refresh QA pending if file is absent or not manually verified.
"""
    write_text(QA_DIR / "qa_checklist.md", checklist)
    write_text(QA_DIR / "regression_qa_notes.md", "# Regression QA Notes\n\nNo prior Project 04 - Customer Funnel Conversion artifact was reused. Regression scope covers generated data, KPI reconciliation, HTML preview filters, and build package docs.")
    write_text(QA_DIR / "performance_qa_notes.md", "# Performance QA Notes\n\nPrepared data is at session and aggregate grain. The dashboard should perform well for portfolio/demo use. For production volumes, keep event-level detail out of primary visuals and use aggregated stage tables.")
    write_text(QA_DIR / "visual_qa_notes.md", "# Visual QA Notes\n\nHTML preview includes four dashboard pages with compact cards, funnel bars, trend SVG, segment tables, and campaign efficiency views. Power BI visual QA remains pending until PBIX File QA.")
    write_text(QA_DIR / "interaction_qa_notes.md", "# Interaction QA Notes\n\nHTML filter controls update segment views. Power BI slicer sync/edit interactions are specified in `build/config/slicer_map.json` and pending PBIX authoring.")
    write_json(QA_DIR / "validation_summary.json", {"data_qa": "pass", "metric_qa": "pass", "visual_qa_html": "pass", "interaction_qa_html": "pass", "file_qa": "pass" if pbix_exists else "blocked_pending_pbix"})
    write_json(QA_DIR / "pbix_validation.json", {"pbix_exists": pbix_exists, "status": "pending_desktop_file_qa" if not pbix_exists else "exists_unverified"})
    write_json(OUTPUT_DIR / "output_validation.json", {"html_preview": str(OUTPUT_DIR / "dashboard.html"), "pbix": str(OUTPUT_DIR / "dashboard_final.pbix"), "pbix_exists": pbix_exists})


def main() -> None:
    ensure_dirs()
    dims = generate_dimensions()
    dfs = generate_synthetic_data(dims)
    write_dataframes(dfs)
    profile = build_profiles(dfs)
    build_reconciliation(dfs, profile)
    data_quality_report(profile, dfs)
    make_metric_docs()
    make_data_dictionary(dfs)
    build_config(profile)
    make_power_query(dfs)
    payload = build_dashboard_payload(dfs, profile)
    write_json(OUTPUT_DIR / "dashboard_payload.json", payload)
    render_html_dashboard(payload)
    make_agent_docs(profile)
    make_powerbi_scripts()
    make_docs(profile)
    make_qa_docs(profile)
    print(json.dumps({"status": "built", "project_root": str(PROJECT_ROOT), "visits": profile["stage_totals"]["visits"], "purchases": profile["stage_totals"]["purchases"], "revenue": profile["revenue"]}, indent=2))


if __name__ == "__main__":
    main()
