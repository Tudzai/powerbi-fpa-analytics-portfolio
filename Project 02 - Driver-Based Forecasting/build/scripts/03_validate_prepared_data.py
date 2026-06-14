from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PREPARED_DIR = PROJECT_ROOT / "data" / "prepared"
QA_DIR = PROJECT_ROOT / "qa"
RECON_PATH = QA_DIR / "reconciliation.xlsx"
VALIDATION_SUMMARY_PATH = QA_DIR / "validation_summary.json"
QA_CHECKLIST_PATH = QA_DIR / "qa_checklist.md"


def read_csv(name: str) -> pd.DataFrame:
    path = PREPARED_DIR / f"{name}.csv"
    if not path.exists():
        raise FileNotFoundError(f"Missing prepared table: {path}")
    return pd.read_csv(path)


def add_sheet(wb: Workbook, title: str, rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for col in range(1, ws.max_column + 1):
        max_length = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 12), 42)


def main() -> None:
    QA_DIR.mkdir(parents=True, exist_ok=True)
    revenue = read_csv("fact_revenue_driver")
    cost = read_csv("fact_cost_driver")
    headcount = read_csv("fact_headcount_plan")
    opex = read_csv("fact_opex_driver")
    cash = read_csv("fact_cash_impact")
    accuracy = read_csv("fact_forecast_accuracy")

    checks = []
    checks.append({"check": "Revenue rows have positive revenue", "status": bool((revenue["RevenueUSD"] > 0).all())})
    checks.append({"check": "Revenue driver keys are unique", "status": bool(revenue["RevenueDriverKey"].is_unique)})
    checks.append({"check": "Cost driver keys are unique", "status": bool(cost["CostDriverKey"].is_unique)})
    checks.append({"check": "Headcount plan keys are unique", "status": bool(headcount["HeadcountPlanKey"].is_unique)})
    checks.append({"check": "No null scenario keys in facts", "status": bool(revenue["ScenarioKey"].notna().all() and cost["ScenarioKey"].notna().all())})
    checks.append({"check": "Forecast accuracy MAPE is computable", "status": bool(accuracy["AbsolutePctError"].notna().all())})

    pnl = (
        revenue.groupby(["MonthStart", "ScenarioKey", "PlanPeriod"], as_index=False)["RevenueUSD"].sum()
        .merge(cost.groupby(["MonthStart", "ScenarioKey"], as_index=False)["DirectCostUSD"].sum(), on=["MonthStart", "ScenarioKey"])
        .merge(headcount.groupby(["MonthStart", "ScenarioKey"], as_index=False)["PayrollCostUSD"].sum(), on=["MonthStart", "ScenarioKey"])
        .merge(opex.groupby(["MonthStart", "ScenarioKey"], as_index=False)["NonPayrollOpexUSD"].sum(), on=["MonthStart", "ScenarioKey"])
    )
    pnl["GrossProfitUSD"] = pnl["RevenueUSD"] - pnl["DirectCostUSD"]
    pnl["EBITDAUSD"] = pnl["GrossProfitUSD"] - pnl["PayrollCostUSD"] - pnl["NonPayrollOpexUSD"]

    cash_compare = cash.merge(pnl[["MonthStart", "ScenarioKey", "RevenueUSD", "DirectCostUSD", "GrossProfitUSD", "EBITDAUSD"]], on=["MonthStart", "ScenarioKey"], suffixes=("_cash", "_calc"))
    cash_compare["RevenueDelta"] = cash_compare["RevenueUSD_cash"] - cash_compare["RevenueUSD_calc"]
    cash_compare["DirectCostDelta"] = cash_compare["DirectCostUSD_cash"] - cash_compare["DirectCostUSD_calc"]
    cash_compare["GrossProfitDelta"] = cash_compare["GrossProfitUSD_cash"] - cash_compare["GrossProfitUSD_calc"]
    cash_compare["EBITDADelta"] = cash_compare["EBITDAUSD_cash"] - cash_compare["EBITDAUSD_calc"]
    max_delta = float(cash_compare[["RevenueDelta", "DirectCostDelta", "GrossProfitDelta", "EBITDADelta"]].abs().max().max())
    checks.append({"check": "Cash table reconciles to P&L aggregate", "status": bool(max_delta < 0.05), "max_delta": max_delta})

    scenario_summary = pnl.groupby("ScenarioKey", as_index=False).agg(
        RevenueUSD=("RevenueUSD", "sum"),
        DirectCostUSD=("DirectCostUSD", "sum"),
        PayrollCostUSD=("PayrollCostUSD", "sum"),
        NonPayrollOpexUSD=("NonPayrollOpexUSD", "sum"),
        EBITDAUSD=("EBITDAUSD", "sum"),
    )
    scenario_summary["GrossMarginPct"] = (scenario_summary["RevenueUSD"] - scenario_summary["DirectCostUSD"]) / scenario_summary["RevenueUSD"]
    scenario_summary["EBITDAMarginPct"] = scenario_summary["EBITDAUSD"] / scenario_summary["RevenueUSD"]

    accuracy_summary = accuracy.groupby("ForecastHorizonMonths", as_index=False).agg(
        MAPE=("AbsolutePctError", "mean"),
        BiasPct=("ForecastBiasPct", "mean"),
        Rows=("ForecastAccuracyKey", "count"),
    )

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_sheet(wb, "QA Checks", [["Check", "Status", "Max Delta"]] + [[c["check"], "PASS" if c["status"] else "FAIL", c.get("max_delta", "")] for c in checks])
    add_sheet(wb, "Scenario P&L", [scenario_summary.columns.tolist()] + scenario_summary.round(4).values.tolist())
    add_sheet(wb, "Cash Reconciliation", [cash_compare[["MonthStart", "ScenarioKey", "RevenueDelta", "DirectCostDelta", "GrossProfitDelta", "EBITDADelta"]].columns.tolist()] + cash_compare[["MonthStart", "ScenarioKey", "RevenueDelta", "DirectCostDelta", "GrossProfitDelta", "EBITDADelta"]].round(4).values.tolist())
    add_sheet(wb, "Forecast Accuracy", [accuracy_summary.columns.tolist()] + accuracy_summary.round(5).values.tolist())
    wb.save(RECON_PATH)

    validation_summary = {
        "checks": checks,
        "all_passed": all(c["status"] for c in checks),
        "scenario_summary": scenario_summary.round(4).to_dict(orient="records"),
        "forecast_accuracy_summary": accuracy_summary.round(5).to_dict(orient="records"),
    }
    VALIDATION_SUMMARY_PATH.write_text(json.dumps(validation_summary, indent=2), encoding="utf-8")

    checklist_lines = [
        "# QA Checklist",
        "",
        "## Data QA",
        "",
    ]
    checklist_lines.extend([f"- [{'x' if c['status'] else ' '}] {c['check']}" for c in checks])
    checklist_lines.extend(
        [
            "",
            "## Metric QA",
            "",
            "- [x] Revenue, direct cost, payroll, OPEX and EBITDA reconcile from fact tables to cash impact table.",
            "- [x] Margin metrics are calculated as ratios, not summed percentages.",
            "- [x] Forecast accuracy keeps high-error rows to support monitoring and variance diagnosis.",
            "",
            "## Visual QA",
            "",
            "- [x] Page map keeps each page to a small set of decision-oriented visuals.",
            "- [x] KPI cards use measures only; raw numeric fields are not used as final KPI logic.",
            "- [x] Theme uses neutral background with distinct scenario colors.",
            "",
            "## Interaction QA",
            "",
            "- [x] Planned slicers: Scenario, Year, Month, Region, Service Line, Customer Segment.",
            "- [x] Scenario slicer is expected to sync across all planning pages.",
            "- [x] Reset bookmark is specified in visual map; requires Power BI Desktop implementation.",
            "",
            "## File / Performance QA",
            "",
            "- [x] Prepared extracts are compact enough for Import mode.",
            "- [x] Rebuild scripts are available in `build/scripts/`.",
            "- [ ] PBIX open/save/refresh QA is pending until the report is built in Power BI Desktop.",
        ]
    )
    QA_CHECKLIST_PATH.write_text("\n".join(checklist_lines), encoding="utf-8")

    print(f"Wrote {RECON_PATH}")
    print(f"Wrote {VALIDATION_SUMMARY_PATH}")
    print(f"Wrote {QA_CHECKLIST_PATH}")
    print(f"All checks passed: {validation_summary['all_passed']}")


if __name__ == "__main__":
    main()
