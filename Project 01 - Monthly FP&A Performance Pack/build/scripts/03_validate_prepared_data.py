from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
PREPARED_DIR = PROJECT_ROOT / "data" / "prepared"
QA_DIR = PROJECT_ROOT / "qa"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def aggregate(rows: list[dict[str, str]], keys: list[str], measures: list[str]) -> dict[tuple[str, ...], dict[str, float]]:
    grouped: dict[tuple[str, ...], dict[str, float]] = defaultdict(lambda: {m: 0.0 for m in measures})
    for row in rows:
        key = tuple(row[k] for k in keys)
        for measure in measures:
            grouped[key][measure] += float(row.get(measure, 0) or 0)
    return grouped


def map_prepared_to_raw_grain(fact_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    dim_date = {r["DateKey"]: r["MonthStart"] for r in read_csv(PREPARED_DIR / "dim_date.csv")}
    dim_scenario = {r["ScenarioKey"]: r["Scenario"] for r in read_csv(PREPARED_DIR / "dim_scenario.csv")}
    converted: list[dict[str, str]] = []
    for row in fact_rows:
        copy = dict(row)
        copy["MonthStart"] = dim_date[row["DateKey"]]
        copy["Scenario"] = dim_scenario[row["ScenarioKey"]]
        converted.append(copy)
    return converted


def relationship_checks() -> list[dict[str, object]]:
    checks: list[dict[str, object]] = []
    facts = {
        "fact_financials.csv": read_csv(PREPARED_DIR / "fact_financials.csv"),
        "fact_opex_department.csv": read_csv(PREPARED_DIR / "fact_opex_department.csv"),
        "fact_cash.csv": read_csv(PREPARED_DIR / "fact_cash.csv"),
        "fact_bridge.csv": read_csv(PREPARED_DIR / "fact_bridge.csv"),
    }
    dims = {
        "DateKey": {r["DateKey"] for r in read_csv(PREPARED_DIR / "dim_date.csv")},
        "ScenarioKey": {r["ScenarioKey"] for r in read_csv(PREPARED_DIR / "dim_scenario.csv")},
        "BusinessUnitKey": {r["BusinessUnitKey"] for r in read_csv(PREPARED_DIR / "dim_business_unit.csv")},
        "ProductKey": {r["ProductKey"] for r in read_csv(PREPARED_DIR / "dim_product.csv")},
        "RegionKey": {r["RegionKey"] for r in read_csv(PREPARED_DIR / "dim_region.csv")},
        "CustomerKey": {r["CustomerKey"] for r in read_csv(PREPARED_DIR / "dim_customer.csv")},
        "DepartmentKey": {r["DepartmentKey"] for r in read_csv(PREPARED_DIR / "dim_department.csv")},
    }
    for fact_name, rows in facts.items():
        for key, valid_values in dims.items():
            if rows and key in rows[0]:
                missing = sorted({row[key] for row in rows if row[key] not in valid_values})
                checks.append(
                    {
                        "CheckType": "Relationship",
                        "Subject": f"{fact_name}.{key}",
                        "Expected": "All fact keys exist in dimension",
                        "Actual": "Pass" if not missing else f"Missing keys: {missing[:5]}",
                        "Delta": "",
                        "Status": "Pass" if not missing else "Fail",
                    }
                )
    return checks


def write_reconciliation_xlsx(rows: list[dict[str, object]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Reconciliation"
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="F4A261")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = header_fill
    for row in rows:
        ws.append([row[h] for h in headers])
    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)), *(len(str(row[header])) for row in rows)) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 42)
    workbook.save(QA_DIR / "reconciliation.xlsx")
    return True


def main() -> None:
    QA_DIR.mkdir(parents=True, exist_ok=True)
    raw_financials = read_csv(RAW_DIR / "raw_fpa_financials.csv")
    prepared_financials = map_prepared_to_raw_grain(read_csv(PREPARED_DIR / "fact_financials.csv"))

    measures = ["Revenue", "COGS", "GrossMargin", "AllocatedOpex", "EBITDA", "Orders", "CashImpact"]
    raw_grouped = aggregate(raw_financials, ["MonthStart", "Scenario"], measures)
    prepared_grouped = aggregate(prepared_financials, ["MonthStart", "Scenario"], measures)
    reconciliation_rows: list[dict[str, object]] = []

    for key in sorted(raw_grouped):
        for measure in measures:
            raw_value = round(raw_grouped[key][measure], 2)
            prepared_value = round(prepared_grouped[key][measure], 2)
            delta = round(prepared_value - raw_value, 2)
            reconciliation_rows.append(
                {
                    "CheckType": "Metric",
                    "Subject": f"{key[0]} {key[1]} {measure}",
                    "Expected": raw_value,
                    "Actual": prepared_value,
                    "Delta": delta,
                    "Status": "Pass" if abs(delta) <= 0.05 else "Fail",
                }
            )

    formula_failures = 0
    for row in raw_financials:
        gm_delta = round(float(row["Revenue"]) - float(row["COGS"]) - float(row["GrossMargin"]), 2)
        ebitda_delta = round(float(row["GrossMargin"]) - float(row["AllocatedOpex"]) - float(row["EBITDA"]), 2)
        if abs(gm_delta) > 0.05 or abs(ebitda_delta) > 0.05:
            formula_failures += 1
    reconciliation_rows.append(
        {
            "CheckType": "Formula",
            "Subject": "Raw financial formula integrity",
            "Expected": "Revenue - COGS = GrossMargin; GrossMargin - AllocatedOpex = EBITDA",
            "Actual": f"{formula_failures} failing rows",
            "Delta": "",
            "Status": "Pass" if formula_failures == 0 else "Fail",
        }
    )
    reconciliation_rows.extend(relationship_checks())

    fields = ["CheckType", "Subject", "Expected", "Actual", "Delta", "Status"]
    write_csv(QA_DIR / "reconciliation.csv", reconciliation_rows, fields)
    xlsx_written = write_reconciliation_xlsx(reconciliation_rows)

    failed = [row for row in reconciliation_rows if row["Status"] != "Pass"]
    notes = [
        "# Prepared Data Validation",
        "",
        f"- Checks run: {len(reconciliation_rows)}",
        f"- Failed checks: {len(failed)}",
        f"- Reconciliation CSV: `qa/reconciliation.csv`",
        f"- Reconciliation XLSX: {'`qa/reconciliation.xlsx`' if xlsx_written else 'not created; openpyxl unavailable'}",
        "",
        "## Result",
        "",
        "PASS" if not failed else "FAIL",
        "",
    ]
    if failed:
        notes.append("## Failed Checks")
        notes.append("")
        for row in failed[:20]:
            notes.append(f"- {row['Subject']}: {row['Actual']}")
    (QA_DIR / "validation_results.md").write_text("\n".join(notes) + "\n", encoding="utf-8")

    if failed:
        raise SystemExit(f"Validation failed with {len(failed)} failing checks.")
    print(f"Validation passed. XLSX written: {xlsx_written}")


if __name__ == "__main__":
    main()
